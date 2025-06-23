import os
import pandas as pd
import numpy as np
from flask import jsonify, request, render_template, session
import json
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.express as px
import uuid
import warnings
import xlrd
import re
from statsmodels.tsa.seasonal import seasonal_decompose  # type: ignore
import logging
import time
import random
import openpyxl
from werkzeug.utils import secure_filename  # 添加导入secure_filename函数
import plotly

# 定义图表颜色常量
CHART_COLORS = {
    'positive': '#e74c3c',  # 红色 - 上涨
    'negative': '#27ae60',  # 绿色 - 下跌
    'primary': '#3498db',   # 蓝色 - 主要趋势线
    'secondary': '#9b59b6', # 紫色 - 次要趋势线
    'neutral': '#7f8c8d'    # 灰色 - 中性
}

# 抑制所有pandas UserWarning
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

# 创建一个全局字典来跟踪正在进行的分析任务
ongoing_analyses = {}
# 创建全局取消标志
analysis_cancelled = False

def handle_sales_trend(app):
    """注册销售趋势分析相关路由"""
    app.add_url_rule('/sales_trend', 'sales_trend_page', sales_trend_page)
    app.add_url_rule('/api/analyze_sales_trend', 'analyze_sales_trend', analyze_sales_trend, methods=['POST'])
    app.add_url_rule('/api/get_analysis_suggestions', 'get_analysis_suggestions', get_analysis_suggestions, methods=['POST'])
    app.add_url_rule('/api/check_year_over_year_eligibility', 'check_year_over_year_eligibility', check_year_over_year_eligibility, methods=['POST'])
    app.add_url_rule('/api/cancel_analysis', 'cancel_analysis', cancel_analysis, methods=['POST'])
    return app

def sales_trend_page():
    """渲染销售趋势分析页面"""
    return render_template('sales_trend.html')

def cancel_analysis():
    """处理取消分析请求"""
    try:
        data = request.json
        analysis_type = data.get('analysis_type')
        operation = data.get('operation', 'analysis')  # 可能的值: 'analysis', 'file_preprocessing'
        timestamp = data.get('timestamp')
        
        # 记录取消请求
        print(f"收到取消请求: 类型={analysis_type}, 操作={operation}, 时间戳={timestamp}")
        
        # 设置全局取消标志
        global analysis_cancelled
        analysis_cancelled = True
        
        # 获取用户名或IP地址作为标识
        username = session.get('username', '匿名用户')
        client_ip = request.remote_addr
        
        # 针对特定用户的所有分析任务进行取消标记
        cancel_count = 0
        for task_id, task_info in list(ongoing_analyses.items()):
            if task_info.get('username') == username or task_info.get('ip') == client_ip:
                task_info['cancelled'] = True
                cancel_count += 1
                print(f"已将任务 {task_id} 标记为已取消（用户：{username}, IP: {client_ip}）")
        
        if cancel_count > 0:
            print(f"已取消 {cancel_count} 个正在进行的任务")
        else:
            print(f"未找到正在进行的任务，但已设置全局取消标志")
        
        # 返回成功响应
        return jsonify({
            'success': True,
            'message': '已成功取消请求',
            'cancelled': True,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"取消请求时出错: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'取消请求时出错: {str(e)}'
        })

def analyze_sales_trend():
    """处理销售趋势分析请求"""
    # 每次开始新的分析，重置取消标志
    global analysis_cancelled
    analysis_cancelled = False
    
    # 创建唯一的任务ID
    task_id = str(uuid.uuid4())
    
    start_time = time.time() 

    try:
        # 记录任务信息
        username = session.get('username', '匿名用户')
        client_ip = request.remote_addr
        
        ongoing_analyses[task_id] = {
            'cancelled': False,
            'start_time': datetime.now(),
            'username': username,
            'ip': client_ip,
            'analysis_type': request.form.get('analysis_type', 'trend')
        }
        
        # 检查是否有文件上传
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '请上传销售数据文件'})
        
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'})
        
        # 获取请求参数
        date_column = request.form.get('date_column', '')
        value_column = request.form.get('value_column', '')
        analysis_type = request.form.get('analysis_type', 'trend')  # 趋势、同比、环比
        time_granularity = request.form.get('time_granularity', 'day')  # 天、周、月、季度、年
        sheet_name = request.form.get('sheet_name', 0)  # 默认使用第一个工作表
        
        if not date_column or not value_column:
            return jsonify({'success': False, 'message': '请选择日期列和值列'})
        
        # 保存文件
        temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 
                              f"temp_{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}")
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        try:
            # 检查是否已取消 - 检查全局标志和特定任务标志
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
                
            # 读取Excel文件，使用指定的工作表
            df = pd.read_excel(temp_path, sheet_name=sheet_name)
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
            
            # 检查列是否存在
            if date_column not in df.columns:
                return jsonify({'success': False, 'message': f'日期列 {date_column} 不存在'})
            if value_column not in df.columns:
                return jsonify({'success': False, 'message': f'值列 {value_column} 不存在'})
            
            # 确保日期列是日期类型
            try:
                # 修改日期解析，增强对多种格式的支持
                # 首先检查是否是数字形式的日期（如20250109）
                sample_dates = df[date_column].head(10).astype(str)
                has_numeric_dates = sample_dates.str.match(r'^\d{8}$').any()
                has_chinese_dates = sample_dates.str.contains(r'年.*月.*日').any()
                
                # 预处理特殊格式
                if has_numeric_dates or has_chinese_dates:
                    # 创建日期转换函数
                    def preprocess_date(date_str):
                        if pd.isna(date_str):
                            return date_str
                            
                        date_str = str(date_str)
                        # 处理数字形式日期
                        if re.match(r'^\d{8}$', date_str):
                            return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                        
                        # 处理中文日期格式
                        match = re.match(r'(\d+)年(\d+)月(\d+)日', date_str)
                        if match:
                            year, month, day = match.groups()
                            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            
                        # 其他格式不变
                        return date_str
                
                    # 应用预处理
                    df[date_column] = df[date_column].astype(str).apply(preprocess_date)
                    
                    #if has_numeric_dates:
                        #print("已处理数字形式的日期格式（如20250109）")
                    #if has_chinese_dates:
                        #print("已处理中文日期格式（如2025年01月09日）")
                
                # 检查是否已取消
                if is_analysis_cancelled(task_id):
                    return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
                
                # 尝试转换为日期时间类型，提供多种格式支持
                df[date_column] = pd.to_datetime(
                    df[date_column], 
                    errors='coerce',            # 无法解析的值设为NaT
                    format=None,                # 自动推断格式
                    dayfirst=False              # 默认月/日/年格式，根据数据实际情况可调整
                )
                
                # 检查并处理解析后的NaT值
                nat_count = df[date_column].isna().sum()
                if nat_count > 0:
                    print(f"警告: {nat_count}行日期值无法解析")
                    
                # 删除NaT值或填充（根据需要选择）
                df = df.dropna(subset=[date_column])
                
                if len(df) == 0:
                    return jsonify({'success': False, 'message': f'处理后没有有效的日期数据'})
            except Exception as e:
                return jsonify({'success': False, 'message': f'无法将 {date_column} 转换为日期格式: {str(e)}'})
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
                
            # 确保值列是数值类型
            try:
                # 先尝试直接转换为数值类型
                try:
                    df[value_column] = pd.to_numeric(df[value_column], errors='raise')
                except:
                    # 如果直接转换失败，尝试处理可能的文本格式数字
                    print(f"将 {value_column} 列从文本格式转换为数值格式")
                    
                    # 创建转换函数，处理常见的非标准数值格式
                    def convert_to_numeric(val):
                        if pd.isna(val):
                            return np.nan
                        try:
                            # 如果已经是数值类型，直接返回
                            if isinstance(val, (int, float)):
                                return val
                                
                            # 处理字符串类型
                            val_str = str(val)
                            
                            # 移除货币符号、千分位分隔符等
                            val_str = val_str.replace(',', '')  # 去掉千分位分隔符
                            val_str = re.sub(r'[$¥€£]', '', val_str)  # 去掉货币符号
                            val_str = val_str.strip()  # 去掉前后空格
                            
                            # 处理百分比
                            if '%' in val_str:
                                val_str = val_str.replace('%', '')
                                return float(val_str) / 100
                                
                            # 转换为浮点数
                            return float(val_str)
                        except:
                            return np.nan
                    
                    # 应用转换函数
                    df[value_column] = df[value_column].apply(convert_to_numeric)
                    
                    # 再次检查转换结果
                    non_na_count = df[value_column].notna().sum()
                    if non_na_count == 0:
                        return jsonify({'success': False, 'message': f'无法将 {value_column} 转换为数值格式，所有值均无效'})
                    
                    # 检查无效值比例
                    na_ratio = 1 - non_na_count / len(df)
                    if na_ratio > 0.5:  # 如果超过50%的值无效
                        print(f"警告: {value_column} 列转换后有 {na_ratio:.2%} 的值为无效值")
                
                # 过滤掉无效的数值行
                df = df.dropna(subset=[value_column])
                if len(df) == 0:
                    return jsonify({'success': False, 'message': f'处理后没有有效的数值数据'})
                    
            except Exception as e:
                return jsonify({'success': False, 'message': f'无法将 {value_column} 转换为数值格式: {str(e)}'})
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
            
            # 对大数据集进行优化
            df = optimize_large_dataframe(df, date_column, value_column)
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
            
            # 根据时间粒度聚合数据
            df = aggregate_by_time(df, date_column, value_column, time_granularity)
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
            
            # 根据分析类型进行相应分析
            if analysis_type == 'trend':
                result = analyze_trend(df, date_column, value_column)
            elif analysis_type == 'year_over_year':
                result = analyze_year_over_year(df, date_column, value_column, time_granularity)
            elif analysis_type == 'month_over_month':
                result = analyze_month_over_month(df, date_column, value_column, time_granularity)
            else:
                return jsonify({'success': False, 'message': f'不支持的分析类型: {analysis_type}'})
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
                
            # 检测异常点
            anomalies = detect_anomalies(df, date_column, value_column, time_granularity=time_granularity)
            
            # 合并结果
            result['anomalies'] = anomalies
            
            # 检查是否已取消
            if is_analysis_cancelled(task_id):
                return jsonify({'success': False, 'message': '分析已被用户取消', 'cancelled': True})
            
            processing_time = time.time() - start_time 

            # 记录分析行为
            try:
                from app import get_db, logger
                # 记录分析操作
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS sales_trend_records (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            username TEXT NOT NULL,
                            file_name TEXT NOT NULL,
                            analysis_type TEXT NOT NULL,
                            time_granularity TEXT NOT NULL,
                            processing_time REAL DEFAULT 0, 
                            created_at DATETIME NOT NULL
                        )
                    ''')
                    cursor.execute("PRAGMA table_info(sales_trend_records)")
                    columns = [column[1] for column in cursor.fetchall()]
                    has_processing_time = 'processing_time' in columns

                    cursor.execute('''
                        INSERT INTO sales_trend_records (
                            username, file_name, analysis_type, time_granularity, processing_time, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        username, 
                        file.filename,
                        analysis_type,
                        time_granularity,
                        processing_time,
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ))
                    conn.commit()
                    logger.info(f'记录销售趋势分析 - 用户: {username}, 文件: {file.filename}')
            except Exception as e:
                print(f"记录分析操作时出错: {str(e)}")
            
            # 清理分析任务状态
            if task_id in ongoing_analyses:
                del ongoing_analyses[task_id]
            
            return jsonify({
                'success': True,
                'message': '分析完成',
                'data': result
            })
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'分析过程中出错: {str(e)}'})
            
        finally:
            # 清理临时文件
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            
            # 清理分析任务状态
            if task_id in ongoing_analyses:
                del ongoing_analyses[task_id]
    
    except Exception as e:
        # 确保清理任务状态
        if task_id in ongoing_analyses:
            del ongoing_analyses[task_id]
        return jsonify({'success': False, 'message': f'处理请求时出错: {str(e)}'})

def is_analysis_cancelled(task_id):
    """检查分析任务是否已被取消"""
    # 检查全局取消标志
    if analysis_cancelled:
        return True
        
    # 检查特定任务的取消标志
    if task_id in ongoing_analyses and ongoing_analyses[task_id].get('cancelled', False):
        return True
        
    return False

def secure_filename(filename):
    """安全化文件名"""
    # 简单实现，实际应用中可以使用werkzeug.utils.secure_filename
    return filename.replace(' ', '_').replace('/', '_')

def aggregate_by_time(df, date_col, value_col, granularity):
    """根据时间粒度聚合数据"""
    # 确保日期列是日期类型
    if not pd.api.types.is_datetime64_any_dtype(df[date_col]):
        #print(f"转换日期列 {date_col} 为日期时间类型")
        # 检查特殊日期格式
        sample_dates = df[date_col].head(10).astype(str)
        has_numeric_dates = sample_dates.str.match(r'^\d{8}$').any()
        has_chinese_dates = sample_dates.str.contains(r'年.*月.*日').any()
        
        # 预处理特殊格式
        if has_numeric_dates or has_chinese_dates:
            # 创建日期转换函数
            def preprocess_date(date_str):
                if pd.isna(date_str):
                    return date_str
                    
                date_str = str(date_str)
                # 处理数字形式日期
                if re.match(r'^\d{8}$', date_str):
                    return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                
                # 处理中文日期格式
                match = re.match(r'(\d+)年(\d+)月(\d+)日', date_str)
                if match:
                    year, month, day = match.groups()
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                    
                # 其他格式不变
                return date_str
                
            # 应用预处理
            df[date_col] = df[date_col].astype(str).apply(preprocess_date)
            
            if has_numeric_dates:
                print("已处理数字形式的日期格式（如20250109）")
            if has_chinese_dates:
                print("已处理中文日期格式（如2025年01月09日）")
        
        # 转换为日期时间
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        # 删除无效日期
        invalid_dates = df[date_col].isna().sum()
        if invalid_dates > 0:
            print(f"警告: 删除 {invalid_dates} 行无效日期")
            df = df.dropna(subset=[date_col])
    
    # 设置日期为索引
    df_copy = df.copy()
    
    # 确保设置索引前日期列没有NaT
    df_copy = df_copy.dropna(subset=[date_col])
    
    # 设置日期列为索引，用于时间序列分析
    df_copy.set_index(date_col, inplace=True)
    
    # 根据不同的时间粒度重采样
    if granularity == 'day':
        df_agg = df_copy.resample('D').sum()
        
        # 日粒度使用精确日期匹配标记原始数据点
        original_dates = set(pd.to_datetime(df[date_col]).dt.strftime('%Y-%m-%d'))
        df_agg = df_agg.reset_index()
        df_agg['is_original_data'] = df_agg[date_col].dt.strftime('%Y-%m-%d').isin(original_dates)
        
    elif granularity == 'week':
        df_agg = df_copy.resample('W-MON').sum()
        
        # 周粒度：只要该周内有任何原始数据，就将整周标记为原始数据
        df_agg = df_agg.reset_index()
        
        # 获取每个原始数据点所在的周
        original_weeks = set()
        for date in pd.to_datetime(df[date_col]):
            # 获取日期所在周的周一
            week_start = date - pd.Timedelta(days=date.dayofweek)
            week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
            original_weeks.add(week_start)
            
        # 标记含有原始数据的周
        df_agg['is_original_data'] = df_agg[date_col].isin(original_weeks)
        print(f"周粒度: 检测到{len(original_weeks)}个含有原始数据的周，总周数为{len(df_agg)}")
        
    elif granularity == 'month':
        df_agg = df_copy.resample('MS').sum()
        
        # 月粒度：只要该月内有任何原始数据，就将整月标记为原始数据
        df_agg = df_agg.reset_index()
        
        # 获取每个原始数据点所在的月
        original_months = set()
        for date in pd.to_datetime(df[date_col]):
            # 获取日期所在月的第一天
            month_start = date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            original_months.add(month_start)
            
        # 标记含有原始数据的月
        df_agg['is_original_data'] = df_agg[date_col].isin(original_months)
        print(f"月粒度: 检测到{len(original_months)}个含有原始数据的月，总月数为{len(df_agg)}")
        
    elif granularity == 'quarter':
        df_agg = df_copy.resample('QS').sum()
        
        # 季度粒度：只要该季度内有任何原始数据，就将整季度标记为原始数据
        df_agg = df_agg.reset_index()
        
        # 获取每个原始数据点所在的季度
        original_quarters = set()
        for date in pd.to_datetime(df[date_col]):
            # 计算季度开始月份 (1, 4, 7, 10)
            quarter_month = (date.month - 1) // 3 * 3 + 1
            # 获取季度的第一天
            quarter_start = date.replace(month=quarter_month, day=1, hour=0, minute=0, second=0, microsecond=0)
            original_quarters.add(quarter_start)
            
        # 标记含有原始数据的季度
        df_agg['is_original_data'] = df_agg[date_col].isin(original_quarters)
        print(f"季度粒度: 检测到{len(original_quarters)}个含有原始数据的季度，总季度数为{len(df_agg)}")
        
    elif granularity == 'year':
        df_agg = df_copy.resample('YS').sum()
        
        # 年粒度：只要该年内有任何原始数据，就将整年标记为原始数据
        df_agg = df_agg.reset_index()
        
        # 获取每个原始数据点所在的年
        original_years = set()
        for date in pd.to_datetime(df[date_col]):
            # 获取年份的第一天
            year_start = date.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            original_years.add(year_start)
            
        # 标记含有原始数据的年
        df_agg['is_original_data'] = df_agg[date_col].isin(original_years)
        print(f"年粒度: 检测到{len(original_years)}个含有原始数据的年，总年数为{len(df_agg)}")
        
    else:
        df_agg = df_copy.resample('D').sum()
        df_agg = df_agg.reset_index()
        # 默认情况下使用日粒度的匹配策略
        original_dates = set(pd.to_datetime(df[date_col]).dt.strftime('%Y-%m-%d'))
        df_agg['is_original_data'] = df_agg[date_col].dt.strftime('%Y-%m-%d').isin(original_dates)
    
    # 绘图需要连续的数据点，仍然填充NaN值，但后续分析时会区分原始数据
    df_agg[value_col] = df_agg[value_col].fillna(0)
    
    return df_agg

def analyze_trend(df, date_col, value_col):
    """分析销售趋势"""
    # 检查数据量级
    row_count = len(df)
    
    # 大数据集性能优化
    if row_count > 100000:
        # 对大数据集进行降采样
        print(f"正在处理大数据集: {row_count}行，将进行降采样")
        
        # 确保日期列格式统一
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        # 根据时间区间降采样，保留关键趋势
        date_min = df[date_col].min()
        date_max = df[date_col].max()
        date_range = (date_max - date_min).days
        
        # 按区间数量决定样本大小
        if date_range > 1000:
            # 对于超过3年的数据，降采样到每周一个点
            df = df.set_index(date_col).resample('W').sum().reset_index()
        elif date_range > 365:
            # 对于超过1年的数据，降采样到每3天一个点
            df = df.set_index(date_col).resample('3D').sum().reset_index()
    
    # 检测当前时间粒度
    # 基于相邻数据点的日期差异来推断时间粒度
    if len(df) >= 2:
        df[date_col] = pd.to_datetime(df[date_col])
        date_diffs = []
        for i in range(1, min(len(df), 10)):  # 检查前10个点或所有点
            diff = (df[date_col].iloc[i] - df[date_col].iloc[i-1]).days
            if diff > 0:  # 避免重复日期
                date_diffs.append(diff)
        
        # 根据平均日期差异判断粒度
        if date_diffs:
            avg_diff = sum(date_diffs) / len(date_diffs)
            if avg_diff >= 90:  # 季度或年
                is_daily_granularity = False
                if avg_diff >= 300:  # 年
                    granularity_type = "年"
                    hover_date_format = '%Y年'
                else:  # 季度
                    granularity_type = "季度"
                    hover_date_format = '%Y年Q%q季度'  # 自定义格式，需要后续处理
            elif avg_diff >= 28:  # 月
                is_daily_granularity = False
                granularity_type = "月"
                hover_date_format = '%Y年%m月'
            elif avg_diff >= 7:  # 周
                is_daily_granularity = False
                granularity_type = "周"
                hover_date_format = '%Y年%m月%d日 (第%U周)'  # %U是一年中的第几周
            else:  # 日
                is_daily_granularity = True
                granularity_type = "日"
                hover_date_format = '%Y年%m月%d日'
        else:
            is_daily_granularity = True  # 默认日粒度
            granularity_type = "默认(日)"
            hover_date_format = '%Y年%m月%d日'
    else:
        is_daily_granularity = True  # 默认日粒度
        granularity_type = "默认(日)"
        hover_date_format = '%Y年%m月%d日'
    
    #print(f"检测到的时间粒度: {granularity_type}")
    
    # 处理季度格式 - Plotly不直接支持季度格式，需要自定义
    if granularity_type == "季度":
        # 添加季度信息
        df['quarter'] = df[date_col].dt.quarter
        # 构建季度显示文本
        df['date_display'] = df[date_col].dt.strftime('%Y年') + 'Q' + df['quarter'].astype(str) + '季度'
    
    # 创建时间序列趋势图 - 根据是否有标记创建不同的图表
    if 'is_original_data' in df.columns:
        # 创建两个数据集 - 一个用于连续线条(所有点)，一个用于标记实际数据点
        df_line = df.copy()
        df_markers = df[df['is_original_data'] == True].copy()
        
        # 创建基础连续线条图(不带悬停信息)
        fig = px.line(df_line, x=date_col, y=value_col, title=f'{value_col}随时间的变化趋势')
        fig.update_traces(line=dict(color=CHART_COLORS['primary'], width=2))
        
        # 根据时间粒度设置合适的悬停模板
        if granularity_type == "季度":
            hover_template = f'日期: %{{customdata}}<br>{value_col}: %{{y}}<extra></extra>'
            hover_data = df_markers['date_display'].tolist()
        else:
            hover_template = f'日期: %{{x|{hover_date_format}}}<br>{value_col}: %{{y}}<extra></extra>'
            hover_data = None
        
        # 添加只有原始数据点的散点，启用悬停信息
        scatter_trace = go.Scatter(
            x=df_markers[date_col],
            y=df_markers[value_col],
            mode='markers',
            name='实际数据点',
            hovertemplate=hover_template,
            marker=dict(size=8, opacity=0.7, color=CHART_COLORS['secondary'])
        )
        
        # 如果是季度格式，添加自定义数据
        if granularity_type == "季度":
            scatter_trace.customdata = hover_data
            
        fig.add_trace(scatter_trace)
        
        # 修改原始线条的悬停设置，隐藏悬停信息
        fig.data[0].hoverinfo = 'skip'
        fig.data[0].name = '趋势线'
    else:
        # 如果没有标记，使用标准图表
        if granularity_type == "季度":
            fig = px.line(df, x=date_col, y=value_col, 
                        title=f'{value_col}随时间的变化趋势', 
                        custom_data=['date_display'])
            fig.update_traces(hovertemplate=f'日期: %{{customdata}}<br>{value_col}: %{{y}}<extra></extra>',
                             line=dict(color=CHART_COLORS['primary'], width=2))
        else:
            fig = px.line(df, x=date_col, y=value_col, 
                        title=f'{value_col}随时间的变化趋势')
            fig.update_traces(hovertemplate=f'日期: %{{x|{hover_date_format}}}<br>{value_col}: %{{y}}<extra></extra>',
                             line=dict(color=CHART_COLORS['primary'], width=2))
    
    fig.update_layout(
        xaxis_title='日期',
        yaxis_title=value_col,
        template='plotly_white',
        hovermode='closest'
    )
    
    # 格式化X轴刻度显示
    if granularity_type == "年":
        fig.update_xaxes(tickformat='%Y年')
    elif granularity_type == "季度":
        # 季度需要自定义刻度
        unique_dates = df[date_col].dt.strftime('%Y年') + 'Q' + df[date_col].dt.quarter.astype(str)
        fig.update_xaxes(tickmode='array', tickvals=df[date_col], ticktext=unique_dates)
    elif granularity_type == "月":
        fig.update_xaxes(tickformat='%Y年%m月')
    elif granularity_type == "周":
        # 周显示月/日
        fig.update_xaxes(tickformat='%m/%d')
    else:  # 日
        # 根据数据范围动态调整格式
        date_range_days = (df[date_col].max() - df[date_col].min()).days
        if date_range_days > 365:
            fig.update_xaxes(tickformat='%Y年%m月')
        elif date_range_days > 60:
            fig.update_xaxes(tickformat='%m月%d日')
        else:
            fig.update_xaxes(tickformat='%m/%d')
    
    # 如果数据点太多，优化图表显示
    if row_count > 1000:
        # 限制显示的刻度数量
        fig.update_xaxes(nticks=20)
        # 增加标记点的间隔
        if 'is_original_data' not in df.columns:
            fig.update_traces(marker=dict(size=3))
    
    # 统计数据计算策略
    # 非日粒度下，将使用所有数据点计算总值，避免数据丢失
    if 'is_original_data' in df.columns:
        # 使用原始数据点
        original_df = df[df['is_original_data'] == True].copy()
        original_count = len(original_df)
        total_count = len(df)
        
        #print(f"原始数据点: {original_count}个，总数据点: {total_count}个")
        
        if original_count > 0:
            if is_daily_granularity:
                # 日粒度时使用原始数据点计算
                #print("使用原始数据点计算统计数据(日粒度)")
                total = original_df[value_col].sum()
                avg = original_df[value_col].mean()
                max_val = original_df[value_col].max()
                max_date = original_df.loc[original_df[value_col].idxmax(), date_col]
                min_val = original_df[value_col].min()
                min_date = original_df.loc[original_df[value_col].idxmin(), date_col]
            else:
                # 非日粒度时使用所有聚合数据点计算
                #print(f"使用所有数据点计算统计数据(非日粒度: {granularity_type})")
                total = df[value_col].sum()
                avg = df[value_col].mean()
                max_val = df[value_col].max()
                max_date = df.loc[df[value_col].idxmax(), date_col]
                min_val = df[value_col].min()
                min_date = df.loc[df[value_col].idxmin(), date_col]
        else:
            # 没有原始数据点的话，使用所有数据点
            #print("没有原始数据点，使用所有数据点计算统计数据")
            total = df[value_col].sum()
            avg = df[value_col].mean()
            max_val = df[value_col].max()
            max_date = df.loc[df[value_col].idxmax(), date_col]
            min_val = df[value_col].min()
            min_date = df.loc[df[value_col].idxmin(), date_col]
    else:
        # 如果没有标记，使用所有数据点
        #print("无数据点标记，使用所有数据点计算统计数据")
        total = df[value_col].sum()
        avg = df[value_col].mean()
        max_val = df[value_col].max()
        max_date = df.loc[df[value_col].idxmax(), date_col]
        min_val = df[value_col].min()
        min_date = df.loc[df[value_col].idxmin(), date_col]
    
    # 根据时间粒度格式化日期显示
    if isinstance(max_date, pd.Timestamp):
        if granularity_type == "年":
            max_date_str = max_date.strftime('%Y年')
            min_date_str = min_date.strftime('%Y年')
        elif granularity_type == "季度":
            max_date_str = f"{max_date.strftime('%Y年')}Q{max_date.quarter}季度"
            min_date_str = f"{min_date.strftime('%Y年')}Q{min_date.quarter}季度"
        elif granularity_type == "月":
            max_date_str = max_date.strftime('%Y年%m月')
            min_date_str = min_date.strftime('%Y年%m月')
        elif granularity_type == "周":
            # 使用format_date_by_granularity函数保持周粒度格式一致性
            # 定义内部函数：与detect_anomalies中的函数保持一致
            def format_date_by_granularity(date_value, granularity):
                if not isinstance(date_value, pd.Timestamp):
                    try:
                        date_value = pd.to_datetime(date_value)
                    except:
                        return str(date_value)  # 无法转换时返回原始字符串
                
                # 根据不同的时间粒度格式化日期
                if granularity == 'day':
                    return date_value.strftime('%Y-%m-%d')
                elif granularity == 'week':
                    # 获取周数 (ISO周)，减去1以符合业务习惯
                    iso_year, iso_week, iso_day = date_value.isocalendar()
                    week_num = iso_week - 1
                    
                    # 计算上一周的日期范围（周一至周日）
                    # 首先计算ISO周的周一日期
                    iso_monday = date_value - pd.Timedelta(days=iso_day-1)
                    
                    # 将日期范围提前一周，然后整体后移一天
                    prev_week_monday = iso_monday - pd.Timedelta(days=6)  # 减少了一天
                    prev_week_sunday = iso_monday  # 现在是本周一，而不是上周日
                    
                    # 根据减1后的周数调整日期和年份
                    if week_num == 0:
                        # 上一年的最后一周
                        prev_year = date_value.year - 1
                        # 上一年的最后一天
                        last_day_prev_year = pd.Timestamp(f"{prev_year}-12-31")
                        last_iso_year, last_iso_week, last_iso_day = last_day_prev_year.isocalendar()
                        week_num = last_iso_week
                        year = prev_year
                    else:
                        year = iso_year
                        
                    # 跨年处理：如果上周的一部分在上一年
                    if prev_week_monday.year < year:
                        # 使用上周一的年份作为显示年份
                        year = prev_week_monday.year
                    
                    # 格式化日期范围
                    date_range = f"{prev_week_monday.month}月{prev_week_monday.day}日-{prev_week_sunday.month}月{prev_week_sunday.day}日"
                    
                    return f"{year}年{week_num}周（{date_range}）"
                elif granularity == 'month':
                    return date_value.strftime('%Y年%m月')
                elif granularity == 'quarter':
                    # 计算季度
                    quarter = (date_value.month - 1) // 3 + 1
                    return f"{date_value.year}年Q{quarter}"
                elif granularity == 'year':
                    return f"{date_value.year}年"
                else:
                    return date_value.strftime('%Y-%m-%d')  # 默认使用日期格式
            
            # 使用format_date_by_granularity函数格式化日期
            max_date_str = format_date_by_granularity(max_date, 'week')
            min_date_str = format_date_by_granularity(min_date, 'week')
        else:  # 日
            max_date_str = max_date.strftime('%Y年%m月%d日')
            min_date_str = min_date.strftime('%Y年%m月%d日')
    else:
        max_date_str = str(max_date)
        min_date_str = str(min_date)
    
    # 计算增长率
    if len(df) >= 2:
        first_val = df[value_col].iloc[0]
        last_val = df[value_col].iloc[-1]
        growth_rate = ((last_val - first_val) / first_val * 100) if first_val != 0 else 0
    else:
        growth_rate = 0
    
    # 尝试进行时间序列分解
    decomposition_result = None
    if len(df) > 6:  # 需要足够的数据点才能进行分解
        try:
            # 设置适当的频率
            if len(df) > 12:
                period = 12  # 假设年度数据，周期为12个月
            else:
                period = len(df) // 2
            
            # 时间序列分解
            decomposition = simple_decompose(df[value_col], period)
            
            # 创建分解趋势图 - 也需要根据时间粒度调整悬停格式
            if granularity_type == "季度" and 'date_display' in df.columns:
                trend_fig = px.line(x=df[date_col], y=decomposition.trend, title='趋势分量', custom_data=df['date_display'])
                trend_fig.update_traces(hovertemplate='日期: %{customdata}<br>值: %{y}<extra></extra>',
                                        line=dict(color=CHART_COLORS['primary'], width=2))
                
                seasonal_fig = px.line(x=df[date_col], y=decomposition.seasonal, title='季节性分量', custom_data=df['date_display'])
                seasonal_fig.update_traces(hovertemplate='日期: %{customdata}<br>值: %{y}<extra></extra>',
                                        line=dict(color=CHART_COLORS['secondary'], width=2))
                
                residual_fig = px.line(x=df[date_col], y=decomposition.resid, title='残差分量', custom_data=df['date_display'])
                residual_fig.update_traces(hovertemplate='日期: %{customdata}<br>值: %{y}<extra></extra>',
                                        line=dict(color='#f39c12', width=2))
            else:
                trend_fig = px.line(x=df[date_col], y=decomposition.trend, title='趋势分量')
                trend_fig.update_traces(hovertemplate=f'日期: %{{x|{hover_date_format}}}<br>值: %{{y}}<extra></extra>',
                                        line=dict(color=CHART_COLORS['primary'], width=2))
                
                seasonal_fig = px.line(x=df[date_col], y=decomposition.seasonal, title='季节性分量')
                seasonal_fig.update_traces(hovertemplate=f'日期: %{{x|{hover_date_format}}}<br>值: %{{y}}<extra></extra>',
                                        line=dict(color=CHART_COLORS['secondary'], width=2))
                
                residual_fig = px.line(x=df[date_col], y=decomposition.resid, title='残差分量')
                residual_fig.update_traces(hovertemplate=f'日期: %{{x|{hover_date_format}}}<br>值: %{{y}}<extra></extra>',
                                        line=dict(color='#f39c12', width=2))
            
            # 同样为分解图设置适当的X轴格式
            for fig_obj in [trend_fig, seasonal_fig, residual_fig]:
                if granularity_type == "年":
                    fig_obj.update_xaxes(tickformat='%Y年')
                elif granularity_type == "季度":
                    # 季度需要自定义刻度
                    fig_obj.update_xaxes(tickmode='array', tickvals=df[date_col], 
                                      ticktext=df[date_col].dt.strftime('%Y年') + 'Q' + df[date_col].dt.quarter.astype(str))
                elif granularity_type == "月":
                    fig_obj.update_xaxes(tickformat='%Y年%m月')
                elif granularity_type == "周":
                    fig_obj.update_xaxes(tickformat='%m/%d')
                else:
                    date_range_days = (df[date_col].max() - df[date_col].min()).days
                    if date_range_days > 365:
                        fig_obj.update_xaxes(tickformat='%Y年%m月')
                    elif date_range_days > 60:
                        fig_obj.update_xaxes(tickformat='%m月%d日')
                    else:
                        fig_obj.update_xaxes(tickformat='%m/%d')
            
            decomposition_result = {
                'trend_chart': trend_fig.to_json(),
                'seasonal_chart': seasonal_fig.to_json(),
                'residual_chart': residual_fig.to_json()
            }
        except Exception as e:
            print(f"时间序列分解失败: {str(e)}")
            decomposition_result = None
    
    return {
        'chart': fig.to_json(),
        'stats': {
            'total': float(total),
            'average': float(avg),
            'max': {
                'value': float(max_val),
                'date': max_date_str
            },
            'min': {
                'value': float(min_val),
                'date': min_date_str
            },
            'growth_rate': float(growth_rate)
        },
        'decomposition': decomposition_result
    }

def analyze_year_over_year(df, date_col, value_col, time_granularity='month'):
    """分析同比增长"""
    # 确保日期列是datetime类型
    df[date_col] = pd.to_datetime(df[date_col])
    
    # 添加年份列
    df['year'] = df[date_col].dt.year
    
    # 根据时间粒度添加适当的时间列
    if time_granularity == 'day':
        df['day_of_year'] = df[date_col].dt.dayofyear
        time_col = 'day_of_year'
        time_label = '日期'
    elif time_granularity == 'week':
        df['week'] = df[date_col].dt.isocalendar().week
        time_col = 'week'
        time_label = '周数'
    elif time_granularity == 'month':
        df['month'] = df[date_col].dt.month
        time_col = 'month'
        time_label = '月份'
    elif time_granularity == 'quarter':
        df['quarter'] = df[date_col].dt.quarter
        time_col = 'quarter'
        time_label = '季度'
    else:  # 'year'
        time_col = 'year'
        time_label = '年份'
    
    # 创建一个空的结果表来存储同比数据
    years = df['year'].unique()
    
    # 只有当时间粒度为"年"时才严格要求至少两年数据
    if time_granularity == 'year' and len(years) < 2:
        return {'success': False, 'message': '数据不足两年，无法进行年度同比分析'}
    
    # 对于其他时间粒度，数据少于两年也可以进行分析，但会给出警告
    if len(years) < 2:
        print(f"警告: 同比分析数据仅包含{len(years)}年，分析结果可能不完整")
    
    # 推断当前时间粒度
    # 根据平均日期差异判断粒度（用于日志和背景信息）
    if len(df) >= 2:
        date_diffs = []
        for i in range(1, min(len(df), 10)):
            diff = (df[date_col].iloc[i] - df[date_col].iloc[i-1]).days
            if diff > 0:
                date_diffs.append(diff)
        
        if date_diffs:
            avg_diff = sum(date_diffs) / len(date_diffs)
            if avg_diff >= 90:
                is_daily_granularity = False
                if avg_diff >= 300:
                    granularity_type = "年"
                    hover_date_format = '%Y年'
                else:
                    granularity_type = "季度"
                    hover_date_format = '%Y年Q%q季度'
            elif avg_diff >= 28:
                is_daily_granularity = False
                granularity_type = "月"
                hover_date_format = '%Y年%m月'
            elif avg_diff >= 7:
                is_daily_granularity = False
                granularity_type = "周"
                hover_date_format = '%Y年%m月%d日 (第%U周)'
            else:
                is_daily_granularity = True
                granularity_type = "日"
                hover_date_format = '%Y年%m月%d日'
        else:
            is_daily_granularity = True
            granularity_type = "默认(日)"
            hover_date_format = '%Y年%m月%d日'
    else:
        is_daily_granularity = True
        granularity_type = "默认(日)"
        hover_date_format = '%Y年%m月%d日'
    
    # 根据时间粒度进行分组
    if time_granularity == 'year':
        # 按年分组，此时不需要额外的时间列
        if 'is_original_data' in df.columns:
            original_df = df[df['is_original_data'] == True].copy()
            if len(original_df) > 0:
                grouped = original_df.groupby(['year'])[value_col].sum().reset_index()
            else:
                grouped = df.groupby(['year'])[value_col].sum().reset_index()
        else:
            grouped = df.groupby(['year'])[value_col].sum().reset_index()
        
        # 重塑数据对于年粒度无需透视
        pivot_df = grouped.set_index('year')[value_col]
        
        # 无需时间名称映射
        time_names = None
    else:
        # 按年和选定的时间粒度分组
        if 'is_original_data' in df.columns and is_daily_granularity:
            original_df = df[df['is_original_data'] == True].copy()
            if len(original_df) > 0:
                grouped = original_df.groupby(['year', time_col])[value_col].sum().reset_index()
            else:
                grouped = df.groupby(['year', time_col])[value_col].sum().reset_index()
        else:
            grouped = df.groupby(['year', time_col])[value_col].sum().reset_index()
        
        # 重塑数据以便于比较
        pivot_df = grouped.pivot(index=time_col, columns='year', values=value_col).fillna(0)
        
        # 创建时间名称映射，确保键为Python原生类型
        if time_granularity == 'month':
            time_names = {
                int(1): "1月", int(2): "2月", int(3): "3月", int(4): "4月", 
                int(5): "5月", int(6): "6月", int(7): "7月", int(8): "8月", 
                int(9): "9月", int(10): "10月", int(11): "11月", int(12): "12月"
            }
        elif time_granularity == 'quarter':
            time_names = {int(1): "Q1", int(2): "Q2", int(3): "Q3", int(4): "Q4"}
        elif time_granularity == 'week':
            # 创建周数映射（1-53周）
            time_names = {int(i): f"{i}周" for i in range(1, 54)}
        elif time_granularity == 'day':
            # 对于天粒度，使用每年的第几天，只需显示数字
            time_names = {int(i): f"{i}天" for i in range(1, 367)}
        else:
            # 默认情况，确保所有键都是Python原生类型
            time_names = {int(i) if isinstance(i, (np.integer, int)) else str(i): str(i) for i in pivot_df.index.tolist()}
    
    # 计算同比增长率
    yoy_changes = {}
    
    if time_granularity == 'year':
        # 年粒度下的增长率计算
        for i in range(1, len(years)):
            year = int(years[i])  # 确保是Python原生int类型
            prev_year = int(years[i-1])
            year_value = float(pivot_df.get(year, 0))
            prev_year_value = float(pivot_df.get(prev_year, 0))
            if prev_year_value != 0:
                yoy_change = (year_value - prev_year_value) / prev_year_value * 100
                yoy_changes[year] = yoy_change
    else:
        # 其他粒度下的增长率计算
        for year in years[1:]:
            year = int(year)  # 确保是Python原生int类型
            prev_year = int(year - 1)
            if prev_year in pivot_df.columns:
                pivot_df[f'{year}_yoy'] = (pivot_df[year] - pivot_df[prev_year]) / pivot_df[prev_year] * 100
                yoy_changes[year] = float(pivot_df[f'{year}_yoy'].mean())
    
    # 创建同比图表
    fig = go.Figure()
    
    # 构建不同年份的颜色列表
    year_colors = [CHART_COLORS['primary'], CHART_COLORS['secondary'], '#2ecc71', '#f39c12', '#9b59b6', '#1abc9c']
    
    # 创建自定义悬停模板
    if time_granularity == 'year':
        hover_template = f'年份: %{{x}}<br>%{{text}}: %{{y:,.2f}}<extra></extra>'
    else:
        hover_template = f'{time_label}: %{{x}}<br>%{{text}}: %{{y:,.2f}}<extra></extra>'
    
    if time_granularity == 'year':
        # 年粒度下的特殊处理
        x_values = [int(year) for year in years]  # 确保是Python原生类型
        y_values = [float(pivot_df.get(int(year), 0)) for year in years]
        hover_texts = [f'{int(year)}年' for year in years]
        
        fig.add_trace(go.Bar(
            x=[f"{year}年" for year in x_values],
            y=y_values,
            text=hover_texts,
            hovertemplate=hover_template,
            marker_color=year_colors[0],
            name=value_col
        ))
        
        # 添加年度增长率
        if yoy_changes:
            growth_years = list(yoy_changes.keys())
            growth_values = list(yoy_changes.values())
            
            fig.add_trace(go.Scatter(
                x=[f"{year}年" for year in growth_years],
                y=growth_values,
                mode='lines+markers',
                name='同比增长率 (%)',
                text=[f'{year}年增长率' for year in growth_years],
                yaxis='y2',
                marker=dict(size=8, color='red'),
                line=dict(color='red', width=2)
            ))
            
            fig.update_layout(
                yaxis2=dict(
                    title='增长率 (%)',
                    overlaying='y',
                    side='right',
                    showgrid=False
                )
            )
    else:
        # 其他粒度的处理
        for i, year in enumerate(years):
            if 'is_original_data' in df.columns and is_daily_granularity:
                # 找出该年份中有实际数据的时间点
                valid_times = set(df[(df['year'] == year) & (df['is_original_data'] == True)][time_col])
                
                # 创建带有自定义文本的数据点
                x_values = pivot_df.index.tolist()
                y_values = pivot_df[year].tolist()
                hover_texts = [f'{year}年' for _ in range(len(x_values))]
                
                # 对于非有效时间点，将悬停信息设置为None
                hover_infos = []
                for time_point in x_values:
                    if time_point in valid_times:
                        hover_infos.append('all')
                    else:
                        hover_infos.append('skip')
                
                # 添加图表
                if time_names:
                    # 使用时间名称映射，确保键被转换为Python原生类型
                    x_labels = [time_names.get(int(t) if isinstance(t, (np.integer, int)) else str(t), str(t)) for t in x_values]
                else:
                    x_labels = [str(t) for t in x_values]
                
                fig.add_trace(go.Scatter(
                    x=x_labels,
                    y=y_values,
                    mode='lines+markers',
                    name=f'{year}年',
                    text=hover_texts,
                    hovertemplate=hover_template,
                    hoverinfo=hover_infos,
                    marker=dict(size=8, color=year_colors[i % len(year_colors)]),
                    line=dict(color=year_colors[i % len(year_colors)], width=2)
                ))
            else:
                # 没有标记或非日粒度，显示所有点
                if time_names:
                    # 使用时间名称映射，确保键被转换为Python原生类型
                    x_labels = [time_names.get(int(t) if isinstance(t, (np.integer, int)) else str(t), str(t)) for t in pivot_df.index.tolist()]
                else:
                    x_labels = [str(t) for t in pivot_df.index.tolist()]
                
                fig.add_trace(go.Scatter(
                    x=x_labels,
                    y=pivot_df[year],
                    mode='lines+markers',
                    name=f'{year}年',
                    text=[f'{year}年' for _ in range(len(pivot_df.index))],
                    hovertemplate=hover_template,
                    marker=dict(size=8, color=year_colors[i % len(year_colors)]),
                    line=dict(color=year_colors[i % len(year_colors)], width=2)
                ))
    
    # 根据时间粒度调整标题和横轴标签
    title_mapping = {
        'day': '日度销售比较',
        'week': '周度销售比较',
        'month': '月度销售比较',
        'quarter': '季度销售比较',
        'year': '年度销售比较'
    }
    
    fig.update_layout(
        title=title_mapping.get(time_granularity, '销售比较'),
        xaxis_title=time_label,
        yaxis_title=value_col,
        template='plotly_white',
        hovermode='closest'
    )
    
    # 创建同比增长率图表
    yoy_fig = go.Figure()
    
    if time_granularity == 'year':
        # 年粒度下不显示单独的增长率图表
        pass
    else:
        for year in years[1:]:
            if f'{year}_yoy' in pivot_df.columns:
                if 'is_original_data' in df.columns and is_daily_granularity:
                    # 找出该年份中有实际数据的时间点
                    valid_times = set(df[(df['year'] == year) & (df['is_original_data'] == True)][time_col])
                    
                    # 创建带有自定义文本的数据点
                    x_values = pivot_df.index.tolist()
                    y_values = pivot_df[f'{year}_yoy'].tolist()
                    hover_texts = [f'{year}年同比' for _ in range(len(x_values))]
                    
                    # 对于非有效时间点，将悬停信息设置为None
                    hover_infos = []
                    for time_point in x_values:
                        if time_point in valid_times:
                            hover_infos.append('all')
                        else:
                            hover_infos.append('skip')
                    
                    # 为正值和负值创建不同的柱状图
                    positive_x = []
                    positive_y = []
                    negative_x = []
                    negative_y = []
                    neutral_x = []
                    neutral_y = []
                    
                    for i, (time_point, val) in enumerate(zip(x_values, y_values)):
                        # 确保time_point是Python原生类型
                        time_point_key = int(time_point) if isinstance(time_point, (np.integer, int)) else str(time_point)
                        time_label_str = time_names.get(time_point_key, str(time_point)) if time_names else str(time_point)
                        if val > 0:
                            positive_x.append(time_label_str)
                            positive_y.append(val)
                        elif val < 0:
                            negative_x.append(time_label_str)
                            negative_y.append(val)
                        else:
                            neutral_x.append(time_label_str)
                            neutral_y.append(val)
                    
                    # 添加柱状图
                    if positive_x:
                        yoy_fig.add_trace(go.Bar(
                            x=positive_x,
                            y=positive_y,
                            name=f'{year}年同比增长',
                            marker_color=CHART_COLORS['positive'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
                    
                    if negative_x:
                        yoy_fig.add_trace(go.Bar(
                            x=negative_x,
                            y=negative_y,
                            name=f'{year}年同比下降',
                            marker_color=CHART_COLORS['negative'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
                    
                    if neutral_x:
                        yoy_fig.add_trace(go.Bar(
                            x=neutral_x,
                            y=neutral_y,
                            name=f'{year}年同比持平',
                            marker_color=CHART_COLORS['neutral'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
                else:
                    # 没有标记或非日粒度，分别创建正值和负值柱状图
                    x_values = pivot_df.index.tolist()
                    y_values = pivot_df[f'{year}_yoy'].tolist()
                    
                    # 为正值和负值创建不同的柱状图
                    positive_x = []
                    positive_y = []
                    negative_x = []
                    negative_y = []
                    neutral_x = []
                    neutral_y = []
                    
                    for time_point, val in zip(x_values, y_values):
                        # 确保time_point是Python原生类型
                        time_point_key = int(time_point) if isinstance(time_point, (np.integer, int)) else str(time_point)
                        time_label_str = time_names.get(time_point_key, str(time_point)) if time_names else str(time_point)
                        if val > 0:
                            positive_x.append(time_label_str)
                            positive_y.append(val)
                        elif val < 0:
                            negative_x.append(time_label_str)
                            negative_y.append(val)
                        else:
                            neutral_x.append(time_label_str)
                            neutral_y.append(val)
                    
                    # 添加柱状图
                    if positive_x:
                        yoy_fig.add_trace(go.Bar(
                            x=positive_x,
                            y=positive_y,
                            name=f'{year}年同比增长',
                            marker_color=CHART_COLORS['positive'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
                    
                    if negative_x:
                        yoy_fig.add_trace(go.Bar(
                            x=negative_x,
                            y=negative_y,
                            name=f'{year}年同比下降',
                            marker_color=CHART_COLORS['negative'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
                    
                    if neutral_x:
                        yoy_fig.add_trace(go.Bar(
                            x=neutral_x,
                            y=neutral_y,
                            name=f'{year}年同比持平',
                            marker_color=CHART_COLORS['neutral'],
                            hovertemplate=f'{time_label}: %{{x}}<br>同比增长率: %{{y:.2f}}%<extra></extra>'
                        ))
        
        growth_title_mapping = {
            'day': '日度同比增长率',
            'week': '周度同比增长率',
            'month': '月度同比增长率',
            'quarter': '季度同比增长率'
        }
        
        yoy_fig.update_layout(
            title=growth_title_mapping.get(time_granularity, '同比增长率'),
            xaxis_title=time_label,
            yaxis_title='同比增长率 (%)',
            template='plotly_white'
        )
    
    # 转换图表为JSON字符串
    chart_json = json.dumps({
        'data': fig.data,
        'layout': fig.layout
    }, cls=plotly.utils.PlotlyJSONEncoder)
    
    if time_granularity == 'year':
        # 年粒度下不需要单独的增长率图表
        yoy_chart_json = None
    else:
        yoy_chart_json = json.dumps({
            'data': yoy_fig.data,
            'layout': yoy_fig.layout
        }, cls=plotly.utils.PlotlyJSONEncoder)
    
    # 计算总计和年度均值
    yearly_totals = {}
    for year in years:
        year = int(year)  # 确保是Python原生int类型
        if time_granularity == 'year':
            yearly_totals[year] = float(pivot_df.get(year, 0))
        else:
            yearly_totals[year] = float(pivot_df[year].sum())
    
    # 准备分析结果
    result = {
        'success': True,
        'chart': chart_json,
        'yoy_chart': yoy_chart_json,
        'stats': {
            'yearly_totals': {str(k): v for k, v in yearly_totals.items()},  # 确保键是字符串
            'yoy_changes': {str(k): float(v) for k, v in yoy_changes.items()}  # 确保键是字符串
        }
    }
    
    if time_granularity != 'year':
        # 为非年粒度准备数据透视表，确保所有键都是字符串
        result['pivot_data'] = {}
        for col in pivot_df.columns:
            col_key = str(col)  # 确保键是字符串
            result['pivot_data'][col_key] = {
                str(idx): float(val) for idx, val in pivot_df[col].items()  # 确保键是字符串
            }
    
    return result

def analyze_month_over_month(df, date_col, value_col, time_granularity='month'):
    """分析环比增长，可根据不同时间粒度进行分析
    
    参数:
    df -- 数据框
    date_col -- 日期列名
    value_col -- 值列名
    time_granularity -- 时间粒度: 'day', 'week', 'month', 'quarter', 'year'
    """
    # 确保日期列是datetime类型
    df[date_col] = pd.to_datetime(df[date_col])
    
    # 根据指定的时间粒度进行分组
    if time_granularity == 'day':
        df['period'] = df[date_col].dt.strftime('%Y-%m-%d')
        df['period_display'] = df[date_col].dt.strftime('%Y年%m月%d日')
        period_title = '日'
        chart_title = '日度销售趋势'
        xaxis_title = '日期'
    elif time_granularity == 'week':
        df['period'] = df[date_col].dt.strftime('%Y-%U')  # ISO周格式
        # 在显示周信息时添加更多细节
        df['week_num'] = df[date_col].dt.isocalendar().week
        
        # 定义周显示格式函数，需要包含日期范围
        def format_week_with_range(date_value):
            if not isinstance(date_value, pd.Timestamp):
                try:
                    date_value = pd.to_datetime(date_value)
                except:
                    return str(date_value)
                    
            # 获取ISO周信息
            iso_year, iso_week, iso_day = date_value.isocalendar()
            # 周数减1，以符合用户习惯
            week_num = iso_week - 1
            
            # 计算该周的周一和周日日期
            start_of_week = date_value - pd.Timedelta(days=iso_day-1)
            end_of_week = start_of_week + pd.Timedelta(days=6)
            
            # 将日期范围往前调整一周
            start_of_week = start_of_week - pd.Timedelta(weeks=1)
            end_of_week = end_of_week - pd.Timedelta(weeks=1)
            
            # 日期范围加一天
            start_of_week = start_of_week + pd.Timedelta(days=1)
            end_of_week = end_of_week + pd.Timedelta(days=1)
            
            # 处理年份交界处的周数
            year_to_display = date_value.year
            if week_num == 0:
                # 如果周数减1后变成0，则应该是上一年的最后一周
                prev_year = year_to_display - 1
                # 获取上一年最后一天的ISO周信息
                last_day_prev_year = pd.Timestamp(f"{prev_year}-12-31")
                _, last_week, _ = last_day_prev_year.isocalendar()
                week_num = last_week
                year_to_display = prev_year
            
            # 格式化成"xxxx年y周（x月x日-x月x日）"的形式
            return f"{year_to_display}年{week_num}周（{start_of_week.month}月{start_of_week.day}日-{end_of_week.month}月{end_of_week.day}日）"
        
        # 使用带日期范围的周格式
        df['period_display'] = df[date_col].apply(format_week_with_range)
        
        period_title = '周'
        chart_title = '周度销售趋势'
        xaxis_title = '周次'
    elif time_granularity == 'month':
        df['period'] = df[date_col].dt.strftime('%Y-%m')
        df['period_display'] = df[date_col].dt.strftime('%Y年%m月')
        period_title = '月'
        chart_title = '月度销售趋势'
        xaxis_title = '年月'
    elif time_granularity == 'quarter':
        df['quarter'] = df[date_col].dt.quarter
        df['period'] = df[date_col].dt.strftime('%Y-') + df['quarter'].astype(str)
        df['period_display'] = df.apply(lambda x: f"{x[date_col].year}年Q{x['quarter']}", axis=1)
        period_title = '季度'
        chart_title = '季度销售趋势'
        xaxis_title = '季度'
    elif time_granularity == 'year':
        df['period'] = df[date_col].dt.strftime('%Y')
        df['period_display'] = df[date_col].dt.strftime('%Y年')
        period_title = '年'
        chart_title = '年度销售趋势'
        xaxis_title = '年份'
    else:
        # 默认使用月份
        df['period'] = df[date_col].dt.strftime('%Y-%m')
        df['period_display'] = df[date_col].dt.strftime('%Y年%m月')
        period_title = '月'
        chart_title = '月度销售趋势'
        xaxis_title = '年月'
    
    #print(f"环比分析 - 使用时间粒度: {time_granularity} ({period_title})")
    
    # 如果数据有is_original_data标记，仅使用标记为原始数据的点进行分析
    if 'is_original_data' in df.columns:
        original_df = df[df['is_original_data'] == True].copy()
        if len(original_df) > 0:
            #print(f"环比分析：使用{len(original_df)}个原始数据点（总点数：{len(df)}）")
            period_data = original_df.groupby(['period', 'period_display'])[value_col].sum().reset_index()
        else:
            # 如果没有原始数据点，使用所有数据
            #print("环比分析：无原始数据点，使用所有数据点")
            period_data = df.groupby(['period', 'period_display'])[value_col].sum().reset_index()
    else:
        # 如果没有标记，使用所有数据点
        print(f"环比分析：使用所有数据点")
        period_data = df.groupby(['period', 'period_display'])[value_col].sum().reset_index()
    
    period_data.sort_values('period', inplace=True)
    
    # 计算环比变化
    period_data['prev_value'] = period_data[value_col].shift(1)
    
    # 添加安全处理，避免除以零
    def safe_pct_change(current, previous):
        if pd.isna(previous) or previous == 0:
            return float('nan')  # 返回NaN而不是触发除以零的警告
        return (current - previous) / previous * 100
    
    # 使用向量化操作安全地计算环比变化率
    period_data['mom_change'] = period_data.apply(
        lambda row: safe_pct_change(row[value_col], row['prev_value']), 
        axis=1
    )
    
    # 创建环比趋势图
    fig = go.Figure()
    
    # 如果数据有is_original_data标记
    if 'is_original_data' in df.columns:
        # 收集原始数据中存在的周期
        original_periods = set(df[df['is_original_data'] == True]['period'])
        
        # 创建所有点的连续线
        fig.add_trace(go.Scatter(
            x=period_data['period_display'],
            y=period_data[value_col],
            mode='lines',
            name='趋势线',
            hoverinfo='skip',
            line=dict(width=2, color=CHART_COLORS['primary'])
        ))
        
        # 仅在原始数据点上添加带悬停信息的标记
        orig_period_data = period_data[period_data['period'].isin(original_periods)]
        
        fig.add_trace(go.Scatter(
            x=orig_period_data['period_display'],
            y=orig_period_data[value_col],
            mode='markers',
            name='实际数据点',
            hovertemplate=f'{period_title}: %{{x}}<br>{value_col}: %{{y:,.2f}}<extra></extra>',
            marker=dict(size=8, opacity=0.7, color=CHART_COLORS['secondary'])
        ))
    else:
        # 没有标记，显示所有点
        fig.add_trace(go.Scatter(
            x=period_data['period_display'],
            y=period_data[value_col],
            mode='lines+markers',
            name=value_col,
            hovertemplate=f'{period_title}: %{{x}}<br>{value_col}: %{{y:,.2f}}<extra></extra>',
            line=dict(width=2, color=CHART_COLORS['primary']),
            marker=dict(size=8, color=CHART_COLORS['primary'])
        ))
    
    fig.update_layout(
        title=chart_title,
        xaxis_title=xaxis_title,
        yaxis_title=value_col,
        template='plotly_white',
        hovermode='closest'
    )
    
    # 创建环比增长率图表
    mom_fig = go.Figure()
    
    if 'is_original_data' in df.columns:
        # 收集原始数据中存在的周期
        original_periods = set(df[df['is_original_data'] == True]['period'])
        
        # 分离正值和负值数据点
        positive_x = []
        positive_y = []
        negative_x = []
        negative_y = []
        neutral_x = []
        neutral_y = []
        
        for i, row in period_data.iterrows():
            # 跳过第一个周期，因为它没有环比数据
            if i == 0 or pd.isna(row['mom_change']):
                continue
                
            period = row['period']
            period_display = row['period_display']
            change = row['mom_change']
            
            # 按环比变化的正负分类
            if change > 0:
                if period in original_periods:
                    positive_x.append(period_display)
                    positive_y.append(change)
            elif change < 0:
                if period in original_periods:
                    negative_x.append(period_display)
                    negative_y.append(change)
            else:
                if period in original_periods:
                    neutral_x.append(period_display)
                    neutral_y.append(change)
        
        # 添加正值柱状图
        if positive_x:
            mom_fig.add_trace(go.Bar(
                x=positive_x,
                y=positive_y,
                name='环比增长',
                marker_color=CHART_COLORS['positive'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
        
        # 添加负值柱状图
        if negative_x:
            mom_fig.add_trace(go.Bar(
                x=negative_x,
                y=negative_y,
                name='环比下降',
                marker_color=CHART_COLORS['negative'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
        
        # 添加零值柱状图
        if neutral_x:
            mom_fig.add_trace(go.Bar(
                x=neutral_x,
                y=neutral_y,
                name='环比持平',
                marker_color=CHART_COLORS['neutral'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
    else:
        # 没有标记，显示所有点，但按环比变化的正负分类
        positive_x = []
        positive_y = []
        negative_x = []
        negative_y = []
        neutral_x = []
        neutral_y = []
        
        for i, row in period_data.iterrows():
            # 跳过第一个周期，因为它没有环比数据
            if i == 0 or pd.isna(row['mom_change']):
                continue
                
            period_display = row['period_display']
            change = row['mom_change']
            
            # 按环比变化的正负分类
            if change > 0:
                positive_x.append(period_display)
                positive_y.append(change)
            elif change < 0:
                negative_x.append(period_display)
                negative_y.append(change)
            else:
                neutral_x.append(period_display)
                neutral_y.append(change)
        
        # 添加正值柱状图
        if positive_x:
            mom_fig.add_trace(go.Bar(
                x=positive_x,
                y=positive_y,
                name='环比增长',
                marker_color=CHART_COLORS['positive'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
        
        # 添加负值柱状图
        if negative_x:
            mom_fig.add_trace(go.Bar(
                x=negative_x,
                y=negative_y,
                name='环比下降',
                marker_color=CHART_COLORS['negative'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
        
        # 添加零值柱状图
        if neutral_x:
            mom_fig.add_trace(go.Bar(
                x=neutral_x,
                y=neutral_y,
                name='环比持平',
                marker_color=CHART_COLORS['neutral'],
                hovertemplate=f'{period_title}: %{{x}}<br>环比增长率: %{{y:.2f}}%<extra></extra>'
            ))
    
    mom_fig.update_layout(
        title=f'{period_title}度环比增长率(%)',
        xaxis_title=xaxis_title,
        yaxis_title='环比增长率(%)',
        template='plotly_white',
        hovermode='closest'
    )
    
    # 汇总统计数据
    positive_changes = (period_data['mom_change'] > 0).sum()
    negative_changes = (period_data['mom_change'] < 0).sum()
    avg_change = period_data['mom_change'].mean()
    
    # 查找最大增长和最大下降的周期
    max_increase_idx = period_data['mom_change'].idxmax()
    max_decrease_idx = period_data['mom_change'].idxmin()
    
    max_increase_period = period_data.loc[max_increase_idx, 'period_display'] if not pd.isna(max_increase_idx) else ""
    max_decrease_period = period_data.loc[max_decrease_idx, 'period_display'] if not pd.isna(max_decrease_idx) else ""
    
    stats = {
        'positive_changes': int(positive_changes),
        'negative_changes': int(negative_changes),
        'average_change': float(avg_change),
        'max_increase': {
            'value': float(period_data['mom_change'].max()),
            'period': max_increase_period
        },
        'max_decrease': {
            'value': float(period_data['mom_change'].min()),
            'period': max_decrease_period
        },
        'period_type': period_title  # 添加周期类型，方便前端显示
    }
    
    # 构建前端友好的周期数据
    period_records = []
    for _, row in period_data.iterrows():
        record = {
            'period': row['period_display'],
            'value': float(row[value_col]),
            'prev_value': float(row['prev_value']) if not pd.isna(row['prev_value']) else None,
            'mom_change': float(row['mom_change']) if not pd.isna(row['mom_change']) else None
        }
        period_records.append(record)
    
    return {
        'chart': fig.to_json(),
        'mom_chart': mom_fig.to_json(),
        'stats': stats,
        'period_data': period_records
    }

def calculate_zscore(series):
    """计算Z分数（标准分数）"""
    mean = np.mean(series)
    std = np.std(series)
    if std == 0:
        return np.zeros(len(series))
    return (series - mean) / std

def calculate_iqr_score(series):
    """计算基于IQR（四分位距）的异常分数"""
    q1 = np.percentile(series, 25)
    q3 = np.percentile(series, 75)
    iqr = q3 - q1
    if iqr == 0:
        return np.zeros(len(series))
    # 标准化为类似Z分数的尺度
    return np.abs(series - np.median(series)) / (iqr / 1.349)  # 1.349 是使IQR与标准差尺度相当的因子

def calculate_mad_score(series):
    """计算基于MAD（中位数绝对偏差）的异常分数"""
    median = np.median(series)
    mad = np.median(np.abs(series - median))
    if mad == 0:
        return np.zeros(len(series))
    # 标准化为类似Z分数的尺度
    return np.abs(series - median) / (mad * 1.4826)  # 1.4826 是使MAD与标准差尺度相当的常数

def detect_consecutive_anomalies(df, date_col, anomaly_markers, window_size=3, min_anomalies=2):
    """
    检测连续时间窗口内的异常累积
    
    参数:
    - df: 数据框
    - date_col: 日期列名
    - anomaly_markers: 异常标记列表 (True/False 或 1/0)
    - window_size: 时间窗口大小，默认为3
    - min_anomalies: 窗口内最小异常数量，默认为2
    
    返回:
    - consecutive_anomalies: 连续异常指标 (0-1之间的值，表示连续性强度)
    - is_in_streak: 是否在异常连续区间内的标记
    """
    df = df.sort_values(by=date_col).reset_index(drop=True)
    n = len(df)
    
    # 初始化结果数组
    consecutive_scores = np.zeros(n)
    is_in_streak = np.zeros(n, dtype=bool)
    
    # 仅在有足够数据时运行
    if n >= window_size:
        #print(f"连续异常检测: 使用{window_size}天窗口，最小异常数{min_anomalies}。强度值含义：")
        #print(f"  - 1.00: 窗口内所有{window_size}个点均为异常 (极高关注度)")
        #print(f"  - {min_anomalies/window_size:.2f}: 窗口内有{min_anomalies}个异常点 (中高关注度)")
        #print(f"  - 0.00: 不在任何连续异常窗口内 (单独异常)")
        
        # 滑动窗口检测连续异常
        total_streaks = 0
        for i in range(n - window_size + 1):
            window = anomaly_markers[i:i+window_size]
            anomaly_count = np.sum(window)
            
            # 如果窗口内异常点数量达到阈值
            if anomaly_count >= min_anomalies:
                total_streaks += 1
                # 计算连续指标 - 窗口内异常点比例
                streak_score = anomaly_count / window_size
                
                # 更新窗口内每个点的连续分数
                for j in range(i, i+window_size):
                    # 如果当前点已标记为异常，增加其连续分数
                    if anomaly_markers[j]:
                        # 多次重叠窗口取最大值
                        consecutive_scores[j] = max(consecutive_scores[j], streak_score)
                        is_in_streak[j] = True
        
        if total_streaks > 0:
            #print(f"检测到{total_streaks}个连续异常窗口")
            # 统计不同强度区间的数量
            strength_bins = {
                "高": np.sum((consecutive_scores > 0.9) & is_in_streak),
                "中高": np.sum((consecutive_scores > 0.6) & (consecutive_scores <= 0.9) & is_in_streak),
                "中": np.sum((consecutive_scores > 0.3) & (consecutive_scores <= 0.6) & is_in_streak)
            }
            for level, count in strength_bins.items():
                if count > 0:
                    #print(f"  - {level}强度连续异常: {count}个点")
                    pass
    
    return consecutive_scores, is_in_streak

def calculate_multidimensional_anomaly_score(series, weights=None):
    """
    计算多维度异常分数
    
    参数:
    - series: 数据序列
    - weights: 各维度权重字典，默认为Z-score 0.5, IQR 0.3, MAD 0.2
    
    返回:
    - combined_scores: 综合异常分数
    - individual_scores: 各维度分数字典
    - threshold_votes: 超过阈值的投票数
    - directions: 异常方向 (1=上升, -1=下降)
    """
    # 填充缺失值为系列均值
    series_filled = series.fillna(series.mean())
    mean_val = series_filled.mean()
    
    # 设置默认权重
    if weights is None:
        weights = {
            "zscore": 0.5,
            "iqr": 0.3,
            "mad": 0.2
        }
    
    # 计算原始Z分数（保留符号方向）
    raw_zscore = (series_filled - mean_val) / series_filled.std()
    
    # 计算各维度分数
    individual_scores = {
        "zscore": np.abs(raw_zscore),  # 只取绝对值用于异常评分
        "iqr": calculate_iqr_score(series_filled),
        "mad": calculate_mad_score(series_filled)
    }
    
    # 计算综合分数
    combined_scores = np.zeros(len(series))
    for score_type, score_values in individual_scores.items():
        combined_scores += weights[score_type] * score_values
    
    # 对异常分数进行投票（至少有两个指标分数超过阈值）
    threshold_votes = np.zeros(len(series), dtype=int)
    for score_values in individual_scores.values():
        threshold_votes += (score_values > 2.5).astype(int)
    
    # 确定每个点的异常方向 (1=上升, -1=下降, 0=不是异常)
    directions = np.zeros(len(series))
    for i in range(len(series)):
        if combined_scores[i] > 2.5 and threshold_votes[i] >= 2:
            # 根据原始Z分数确定方向，保留异常方向信息
            directions[i] = 1 if raw_zscore[i] > 0 else -1
    
    # 返回综合分数、各维度分数、投票结果和方向
    return combined_scores, individual_scores, threshold_votes, directions

def is_holiday(date):
    """判断日期是否为中国主要节假日（基于2018-2030年的精确日期）"""
    if not isinstance(date, pd.Timestamp):
        return False, None, []
    
    # 转换为日期字符串 MM-DD 格式，用于固定日期的节日
    date_str = f"{date.month:02d}-{date.day:02d}"
    year = date.year
    
    # 定义主要节假日的精确日期范围和相关信息
    # 格式: {年份: {节日名称: [日期范围(MM-DD格式)列表]}}
    holiday_dates = {
        # 春节（农历新年）- 具体日期随年份变化
        "春节": {
            2018: ["02-15", "02-16", "02-17", "02-18", "02-19", "02-20", "02-21"],
            2019: ["02-04", "02-05", "02-06", "02-07", "02-08", "02-09", "02-10"],
            2020: ["01-24", "01-25", "01-26", "01-27", "01-28", "01-29", "01-30"],
            2021: ["02-11", "02-12", "02-13", "02-14", "02-15", "02-16", "02-17"],
            2022: ["01-31", "02-01", "02-02", "02-03", "02-04", "02-05", "02-06"],
            2023: ["01-21", "01-22", "01-23", "01-24", "01-25", "01-26", "01-27"],
            2024: ["02-10", "02-11", "02-12", "02-13", "02-14", "02-15", "02-16"],
            2025: ["01-29", "01-30", "01-31", "02-01", "02-02", "02-03", "02-04"],
            2026: ["02-17", "02-18", "02-19", "02-20", "02-21", "02-22", "02-23"],
            2027: ["02-06", "02-07", "02-08", "02-09", "02-10", "02-11", "02-12"],
            2028: ["01-26", "01-27", "01-28", "01-29", "01-30", "01-31", "02-01"],
            2029: ["02-13", "02-14", "02-15", "02-16", "02-17", "02-18", "02-19"],
            2030: ["02-03", "02-04", "02-05", "02-06", "02-07", "02-08", "02-09"],
            "pre_days": 7,  # 春节前7天
            "post_days": 7,  # 春节后7天
            "up_patterns": ["春节前购物高峰", "年货采购期", "春节备货期", "节前消费高峰"],
            "down_patterns": ["春节期间多数商家休息", "春节后初期消费低迷", "节日消费后的淡季", "春节期间商业活动减少"]
        },
        
        # 国庆节黄金周（固定10月1日-7日）
        "国庆节": {
            # 国庆节是固定的每年10月1日至7日
            "fixed_range": ["10-01", "10-02", "10-03", "10-04", "10-05", "10-06", "10-07"],
            "pre_days": 3,
            "post_days": 3,
            "up_patterns": ["国庆节促销活动", "长假旅游消费增加", "黄金周消费高峰", "节日庆祝相关消费"],
            "down_patterns": ["国庆期间部分商家暂停营业", "节日期间商业区客流变化", "供应链受节假日影响"]
        },
        
        # 元旦（固定1月1日）
        "元旦": {
            "fixed_range": ["01-01"],
            "extended_range": ["12-30", "12-31", "01-02", "01-03"],  # 考虑元旦前后
            "pre_days": 2,
            "post_days": 2,
            "up_patterns": ["元旦促销", "年末购物", "新年促销活动", "跨年消费"],
            "down_patterns": ["元旦假期部分商家休息", "新年假期商业活动减少"]
        },
        
        # 劳动节（固定5月1日，但假期长度因年而异）
        "劳动节": {
            2018: ["05-01"],
            2019: ["05-01", "05-02", "05-03", "05-04"],
            2020: ["05-01", "05-02", "05-03", "05-04", "05-05"],
            2021: ["05-01", "05-02", "05-03", "05-04", "05-05"],
            2022: ["04-30", "05-01", "05-02", "05-03", "05-04"],
            2023: ["04-29", "04-30", "05-01", "05-02", "05-03"],
            2024: ["05-01", "05-02", "05-03", "05-04", "05-05"],
            2025: ["05-01", "05-02", "05-03", "05-04", "05-05"],  # 预估
            2026: ["05-01", "05-02", "05-03", "05-04", "05-05"],  # 预估
            2027: ["05-01", "05-02", "05-03", "05-04", "05-05"],  # 预估
            2028: ["04-29", "04-30", "05-01", "05-02", "05-03"],  # 预估
            2029: ["04-28", "04-29", "04-30", "05-01", "05-02"],  # 预估
            2030: ["05-01", "05-02", "05-03", "05-04", "05-05"],  # 预估
            "pre_days": 2,
            "post_days": 2,
            "up_patterns": ["五一促销活动", "小长假消费增加", "假期旅游带动消费", "假日特惠活动"],
            "down_patterns": ["劳动节假期部分商家休息", "假期结束后消费低迷", "商业区域人流变化"]
        },
        
        # 中秋节（农历节日，每年日期不同）
        "中秋节": {
            2018: ["09-24"],
            2019: ["09-13"],
            2020: ["10-01"],  # 与国庆重合
            2021: ["09-21"],
            2022: ["09-10"],
            2023: ["09-29"],
            2024: ["09-17"],
            2025: ["10-06"],  # 与国庆重合
            2026: ["09-25"],
            2027: ["09-15"],
            2028: ["10-03"],  # 与国庆重合
            2029: ["09-22"],
            2030: ["09-12"],
            "pre_days": 3,
            "post_days": 1,
            "up_patterns": ["中秋节礼品销售高峰", "月饼等节日食品销售增加", "中秋团圆消费", "节日礼品采购"],
            "down_patterns": ["中秋节当天销售下降", "节日期间特定商品销量变化"]
        },
        
        # 双十一购物节（固定11月11日）
        "双十一": {
            "fixed_range": ["11-11"],
            "extended_range": ["11-01", "11-02", "11-03", "11-04", "11-05", "11-06", "11-07", "11-08", "11-09", "11-10", "11-12", "11-13", "11-14", "11-15"],
            "pre_days": 10,  # 双11前10天
            "post_days": 5,   # 双11后5天
            "up_patterns": ["双十一购物狂欢节", "大规模促销活动", "预售活动期", "购物节大促"],
            "down_patterns": ["双十一后消费疲软", "透支消费降低后续购买力", "促销后的销售低谷"]
        },
        
        # 双十二（固定12月12日）
        "双十二": {
            "fixed_range": ["12-12"],
            "extended_range": ["12-07", "12-08", "12-09", "12-10", "12-11", "12-13", "12-14", "12-15", "12-16", "12-17"],
            "pre_days": 5,
            "post_days": 3,
            "up_patterns": ["双十二促销活动", "年末购物季", "双十二特惠", "年终促销"],
            "down_patterns": ["双十二后消费下降", "年末消费逐渐减少"]
        }
    }
    
    # 检查是否在特定年份的节假日范围内
    for holiday_name, holiday_info in holiday_dates.items():
        # 检查固定日期范围（如国庆节、元旦等固定日期的节日）
        if "fixed_range" in holiday_info and date_str in holiday_info["fixed_range"]:
            return True, holiday_name, holiday_info["up_patterns"] if date.day <= 3 else holiday_info["down_patterns"]
        
        # 检查扩展日期范围（节日前后几天）
        if "extended_range" in holiday_info and date_str in holiday_info["extended_range"]:
            # 根据日期判断是节前还是节后
            if date.day < 15:  # 简化判断，前半月认为是节前
                return True, f"{holiday_name}前", holiday_info["up_patterns"]
            else:
                return True, f"{holiday_name}后", holiday_info["down_patterns"]
        
        # 检查年份特定的节假日日期（如春节、中秋等农历节日）
        if year in holiday_info and isinstance(holiday_info[year], list) and date_str in holiday_info[year]:
            # 找出该年份该节日的第一天和最后一天
            first_day = holiday_info[year][0]
            last_day = holiday_info[year][-1]
            
            # 判断是节日的前期、中期还是后期
            if date_str == first_day:
                return True, f"{holiday_name}开始", holiday_info["up_patterns"]
            elif date_str == last_day:
                return True, f"{holiday_name}结束", holiday_info["down_patterns"]
            else:
                return True, holiday_name, holiday_info["up_patterns"] if random.random() > 0.5 else holiday_info["down_patterns"]
    
    # 检查节日前后的日期
    date_minus_1 = (date - pd.Timedelta(days=1)).strftime("%m-%d")
    date_minus_2 = (date - pd.Timedelta(days=2)).strftime("%m-%d")
    date_plus_1 = (date + pd.Timedelta(days=1)).strftime("%m-%d")
    date_plus_2 = (date + pd.Timedelta(days=2)).strftime("%m-%d")
    
    for holiday_name, holiday_info in holiday_dates.items():
        # 检查固定日期节日的前后几天
        if "fixed_range" in holiday_info:
            first_day = holiday_info["fixed_range"][0]
            last_day = holiday_info["fixed_range"][-1]
            
            # 检查是否是节日前几天
            if date_plus_1 == first_day or date_plus_2 == first_day:
                return True, f"{holiday_name}前", holiday_info["up_patterns"]
            
            # 检查是否是节日后几天
            if date_minus_1 == last_day or date_minus_2 == last_day:
                return True, f"{holiday_name}后", holiday_info["down_patterns"]
        
        # 检查年份特定节日的前后几天
        if year in holiday_info and isinstance(holiday_info[year], list):
            first_day = holiday_info[year][0]
            last_day = holiday_info[year][-1]
            
            # 检查节日前的日期
            pre_dates = [(pd.Timestamp(f"{year}-{first_day}") - pd.Timedelta(days=i)).strftime("%m-%d") for i in range(1, holiday_info.get("pre_days", 3) + 1)]
            if date_str in pre_dates:
                return True, f"{holiday_name}前", holiday_info["up_patterns"]
            
            # 检查节日后的日期
            post_dates = [(pd.Timestamp(f"{year}-{last_day}") + pd.Timedelta(days=i)).strftime("%m-%d") for i in range(1, holiday_info.get("post_days", 3) + 1)]
            if date_str in post_dates:
                return True, f"{holiday_name}后", holiday_info["down_patterns"]
    
    return False, None, []

def analyze_business_impact(df, idx, date_col, value_col, direction, time_granularity='day'):
    """
    分析销售异常对业务的具体影响，提供业务导向的异常解释
    
    参数:
    - df: 数据框
    - idx: 异常点索引
    - date_col: 日期列名
    - value_col: 值列名
    - direction: 异常方向 ('上升' 或 '下降')
    - time_granularity: 时间粒度，可选值: 'day', 'week', 'month', 'quarter', 'year'
    
    返回:
    - business_reasons: 业务导向的解释列表
    """
    business_reasons = []
    date = df.iloc[idx][date_col]
    value = df.iloc[idx][value_col]
    
    # 根据时间粒度设置相应的时间单位描述和比较窗口
    time_unit_map = {
        'day': '日',
        'week': '周',
        'month': '月',
        'quarter': '季度',
        'year': '年'
    }
    time_unit = time_unit_map.get(time_granularity, '日')
    
    # 根据时间粒度设置比较窗口和描述
    if time_granularity == 'day':
        compare_window = 30
        compare_unit = '天'
    elif time_granularity == 'week':
        compare_window = 8
        compare_unit = '周'
    elif time_granularity == 'month':
        compare_window = 6
        compare_unit = '月'
    elif time_granularity == 'quarter':
        compare_window = 4
        compare_unit = '季度'
    elif time_granularity == 'year':
        compare_window = 3
        compare_unit = '年'
    else:
        compare_window = 30
        compare_unit = '天'
    
    # 确保数据被排序
    try:
        df_sorted = df.sort_values(by=date_col).reset_index(drop=True)
        # 找到当前点在排序后数据中的位置
        current_pos = df_sorted[df_sorted[date_col] == date].index[0]
    except:
        df_sorted = df
        current_pos = idx
    
    # 获取前一个数据点（如果存在）
    prev_value = None
    if current_pos > 0:
        prev_value = df_sorted.iloc[current_pos-1][value_col]
        prev_date = df_sorted.iloc[current_pos-1][date_col]
        
        # 根据时间粒度格式化前一个时间点
        if isinstance(prev_date, pd.Timestamp):
            if time_granularity == 'day':
                prev_time_str = prev_date.strftime('%m-%d')
            elif time_granularity == 'week':
                iso_year, iso_week, _ = prev_date.isocalendar()
                prev_time_str = f"{iso_year}年{iso_week-1}周"
            elif time_granularity == 'month':
                prev_time_str = prev_date.strftime('%Y年%m月')
            elif time_granularity == 'quarter':
                quarter = (prev_date.month - 1) // 3 + 1
                prev_time_str = f"{prev_date.year}年Q{quarter}"
            elif time_granularity == 'year':
                prev_time_str = f"{prev_date.year}年"
            else:
                prev_time_str = prev_date.strftime('%m-%d')
        else:
            prev_time_str = str(prev_date)
    
    # 1. ROI影响分析
    try:
        # 计算与前N个时间单位平均值的偏差百分比
        lookback = compare_window
        start_idx = max(0, current_pos - lookback)
        if current_pos > start_idx:
            prev_avg = df_sorted.iloc[start_idx:current_pos][value_col].mean()
            if prev_avg > 0:
                change_pct = (value - prev_avg) / prev_avg * 100
                
                # 根据变化幅度提供ROI相关解释
                if direction == '上升':
                    if change_pct > 50:
                        business_reasons.append(f"{time_unit}销售额较{compare_window}{compare_unit}均值增长{change_pct:.1f}%，可能大幅提升营销ROI")
                    elif change_pct > 20:
                        business_reasons.append(f"{time_unit}销售额较{compare_window}{compare_unit}均值增长{change_pct:.1f}%，预计改善投资回报率")
                else:  # 下降
                    if change_pct < -40:
                        business_reasons.append(f"{time_unit}销售额较{compare_window}{compare_unit}均值下降{abs(change_pct):.1f}%，可能严重影响营销ROI")
                    elif change_pct < -15:
                        business_reasons.append(f"{time_unit}销售额较{compare_window}{compare_unit}均值减少{abs(change_pct):.1f}%，预计降低投资效益")
                    
            # 同时计算与前一时间单位的变化（更直观的变化）
            if prev_value is not None and prev_value > 0:
                day_change_pct = (value - prev_value) / prev_value * 100
                day_change_abs = abs(day_change_pct)
                
                # 只有变化比较大时才添加到原因中
                if day_change_abs > 15:
                    if day_change_pct > 0:
                        business_reasons.append(f"单{time_unit}销售额较前一{time_unit}({prev_time_str})增长{day_change_pct:.1f}%，环比显著上升")
                    else:
                        business_reasons.append(f"单{time_unit}销售额较前一{time_unit}({prev_time_str})下降{day_change_abs:.1f}%，环比明显减少")
    except Exception as e:
        print(f"ROI影响分析出错: {str(e)}")
    
    # 2. 销售路径分析
    try:
        # 分析销售变化趋势和模式 - 调整窗口大小
        recent_window = max(3, min(7, compare_window // 2))  # 根据比较窗口调整，但至少3个点，最多7个点
        start_recent = max(0, current_pos - recent_window)
        
        if current_pos > start_recent + 2:  # 确保有足够数据进行模式分析
            # 计算与前一周期的环比变化
            recent_values = df_sorted.iloc[start_recent:current_pos+1][value_col].values
            
            # 销售路径分析 - 检测销售漏斗变化模式
            if direction == '上升':
                # 检测是否在持续增长后的爆发
                if len(recent_values) >= 3 and all(recent_values[i] <= recent_values[i+1] for i in range(len(recent_values)-2)):
                    business_reasons.append(f"销售路径转化顺畅，呈连续{recent_window}个{time_unit}的累积上升趋势突破")
                
                # 检测是否是前期投入的延迟转化
                if len(recent_values) >= 5 and recent_values[0] > recent_values[1] and recent_values[-1] > recent_values[-2]:
                    business_reasons.append(f"销售漏斗末端转化率提升，前期{time_unit}营销活动开始显效")
            else:  # 下降
                # 检测是否在持续下降后的崩塌
                if len(recent_values) >= 3 and all(recent_values[i] >= recent_values[i+1] for i in range(len(recent_values)-2)):
                    business_reasons.append(f"销售路径持续恶化，客户流失持续{recent_window}个{time_unit}")
                
                # 检测是否是销售路径早期环节的问题
                if len(recent_values) >= 5 and sum(recent_values[:3]) / 3 < sum(recent_values[-3:]) / 3:
                    business_reasons.append(f"销售漏斗早期环节可能出现问题，新客获取从{recent_window}个{time_unit}前开始减少")
    except Exception as e:
        print(f"销售路径分析出错: {str(e)}")
    
    # 3. 转化率异常分析 - 根据时间粒度调整
    try:
        if isinstance(date, pd.Timestamp):
            # 根据时间粒度提供不同的转化率解释
            if time_granularity == 'day':
                month = date.month
                day = date.day
                weekday = date.dayofweek
                
                # 月度转化周期分析
                if day <= 5:
                    if direction == '上升':
                        business_reasons.append("月初转化率通常较高，新预算释放期")
                    else:
                        business_reasons.append("月初转化下降，可能指示预算分配问题")
                elif day >= 25:
                    if direction == '上升':
                        business_reasons.append("月末转化提升，可能与销售考核周期相关")
                    else:
                        business_reasons.append("月末转化率下降，预算可能已耗尽")
                
                # 周度转化周期分析
                if weekday == 0:  # 周一
                    if direction == '下降':
                        business_reasons.append("周一转化率降低，客户决策周期开始")
                elif weekday == 4:  # 周五
                    if direction == '上升':
                        business_reasons.append("周五转化率提升，周末前决策加速")
                
            elif time_granularity == 'week':
                month = date.month
                week_in_month = (date.day - 1) // 7 + 1
                
                if week_in_month == 1:
                    if direction == '上升':
                        business_reasons.append("月初周转化率提升，新预算周期开始")
                    else:
                        business_reasons.append("月初周转化率下降，可能是预算调整期")
                elif week_in_month >= 4:
                    if direction == '上升':
                        business_reasons.append("月末周转化率飙升，销售冲刺阶段")
                    else:
                        business_reasons.append("月末周转化率下滑，月度预算可能接近耗尽")
                        
            elif time_granularity == 'month':
                month = date.month
                
                # 季度转化周期分析
                if month in [1, 4, 7, 10]:  # 季初月
                    if direction == '上升':
                        business_reasons.append("季度初月转化率上升，新季度计划启动")
                    else:
                        business_reasons.append("季度初月转化率下降，可能是战略调整期")
                elif month in [3, 6, 9, 12]:  # 季末月
                    if direction == '上升':
                        business_reasons.append("季度末月转化率提升，季度目标冲刺")
                    else:
                        business_reasons.append("季度末月转化率下降，季度预算可能已超支")
                        
            elif time_granularity == 'quarter':
                quarter = (date.month - 1) // 3 + 1
                
                if quarter in [1]:  # 年初季度
                    if direction == '上升':
                        business_reasons.append("年初季度转化率提升，年度战略良好开局")
                    else:
                        business_reasons.append("年初季度转化率下滑，年度计划可能需要调整")
                elif quarter in [4]:  # 年末季度
                    if direction == '上升':
                        business_reasons.append("年末季度转化率飙升，年度业绩冲刺")
                    else:
                        business_reasons.append("年末季度转化率下降，可能影响全年业绩")
            
            # 季度转化周期分析（对日粒度的补充）
            if time_granularity == 'day' and date.month in [3, 6, 9, 12] and date.day >= 25:
                if direction == '上升':
                    business_reasons.append("季度末转化率飙升，销售冲刺期")
    except Exception as e:
        print(f"转化率分析出错: {str(e)}")
    
    # 4. 客户行为分析 - 根据时间粒度调整
    try:
        if isinstance(date, pd.Timestamp):
            # 根据时间粒度提供不同的客户行为解释
            if time_granularity in ['day', 'week', 'month']:
                # 尝试从日期模式推断客户行为
                if direction == '上升':
                    if date.month in [11, 12] or (date.month == 1 and date.day <= 15):
                        business_reasons.append(f"年末及新年促销期，高客单价客户{time_unit}活跃度增加")
                    elif date.month in [7, 8]:
                        business_reasons.append(f"暑期促销季，休闲类消费{time_unit}转化率提升")
                else:  # 下降
                    if date.month in [2] and date.day > 10:
                        business_reasons.append(f"春节后消费疲软期，高价值客户{time_unit}活跃度降低")
                    elif date.month in [4, 5]:
                        business_reasons.append(f"传统淡季，客户{time_unit}决策周期延长")
            elif time_granularity == 'quarter':
                if direction == '上升':
                    if date.month in [10, 11, 12]:  # Q4
                        business_reasons.append("第四季度高客单价客户活跃度季节性提升")
                    elif date.month in [4, 5, 6]:  # Q2
                        business_reasons.append("第二季度消费意愿提升，新品上市季")
                else:
                    if date.month in [1, 2, 3]:  # Q1
                        business_reasons.append("第一季度传统淡季，客户季度性消费降低")
                    elif date.month in [7, 8, 9]:  # Q3
                        business_reasons.append("第三季度客户决策延缓，假期效应")
            elif time_granularity == 'year':
                # 年度客户行为趋势
                if direction == '上升':
                    business_reasons.append("年度客户忠诚度提升，复购率增长")
                else:
                    business_reasons.append("年度客户流失率上升，品牌吸引力可能下降")
    except Exception as e:
        print(f"客户行为分析出错: {str(e)}")
        
    return business_reasons

def analyze_multidimensional_correlation(df, date_col, value_col, idx, time_granularity='day'):
    """
    分析多维度数据与销售异常的相关性，提供更全面的异常原因解释
    
    参数:
    - df: 数据框
    - date_col: 日期列名
    - value_col: 销售值列名
    - idx: 异常点索引
    - time_granularity: 时间粒度，可选值: 'day', 'week', 'month', 'quarter', 'year'
    
    返回:
    - correlation_insights: 相关性分析结果列表
    """
    correlation_insights = []
    try:
        # 获取异常点日期和值
        abnormal_date = df.iloc[idx][date_col]
        abnormal_value = df.iloc[idx][value_col]
        
        # 确保日期列是时间类型
        if not pd.api.types.is_datetime64_any_dtype(df[date_col]):
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        # 根据时间粒度设置分析窗口大小
        if time_granularity == 'day':
            lookback_window = 14    # 前14天
            lookforward_window = 7  # 后7天
        elif time_granularity == 'week':
            lookback_window = 8     # 前8周
            lookforward_window = 4  # 后4周
        elif time_granularity == 'month':
            lookback_window = 6     # 前6个月
            lookforward_window = 3  # 后3个月
        elif time_granularity == 'quarter':
            lookback_window = 4     # 前4个季度
            lookforward_window = 2  # 后2个季度
        elif time_granularity == 'year':
            lookback_window = 3     # 前3年
            lookforward_window = 1  # 后1年
        else:
            lookback_window = 14
            lookforward_window = 7
        
        # 确保数据按日期排序
        df_sorted = df.sort_values(by=date_col).reset_index(drop=True)
        
        # 查找异常点在排序后数据中的位置
        current_pos = df_sorted[df_sorted[date_col] == abnormal_date].index[0]
        
        # 设置分析窗口
        start_idx = max(0, current_pos - lookback_window)
        end_idx = min(len(df_sorted) - 1, current_pos + lookforward_window)
        
        # 提取窗口内的数据
        window_df = df_sorted.iloc[start_idx:end_idx+1].copy()
        
        # 检测潜在的相关维度列
        potential_dimensions = []
        
        # 查找数值类型列作为潜在相关维度
        for col in df.columns:
            if col != date_col and col != value_col and pd.api.types.is_numeric_dtype(df[col]):
                potential_dimensions.append(col)
        
        # 如果找到了潜在维度
        if potential_dimensions:
            # 分析每个维度与销售值的相关性
            correlation_results = {}
            
            for dim in potential_dimensions:
                # 计算相关系数
                if window_df[dim].std() > 0 and window_df[value_col].std() > 0:  # 避免常量列
                    # 皮尔逊相关系数(线性关系)
                    pearson_corr = window_df[dim].corr(window_df[value_col], method='pearson')
                    # 斯皮尔曼相关系数(单调关系)
                    spearman_corr = window_df[dim].corr(window_df[value_col], method='spearman')
                    
                    # 判断相关性强度
                    avg_corr = (abs(pearson_corr) + abs(spearman_corr)) / 2
                    
                    # 计算异常点前后该维度的变化
                    if current_pos > start_idx:
                        before_avg = window_df.iloc[:current_pos-start_idx][dim].mean()
                        point_value = window_df.iloc[current_pos-start_idx][dim]
                        
                        if before_avg > 0:  # 避免除零
                            change_pct = (point_value - before_avg) / before_avg * 100
                            
                            # 记录高相关性维度及其变化情况
                            if avg_corr >= 0.5:  # 相关性阈值
                                correlation_results[dim] = {
                                    'correlation': avg_corr,
                                    'change_pct': change_pct,
                                    'pearson_corr': pearson_corr,
                                    'spearman_corr': spearman_corr,
                                    'before_avg': before_avg,
                                    'point_value': point_value
                                }
            
            # 找出最相关的维度(最多3个)
            top_dimensions = sorted(correlation_results.items(), 
                                   key=lambda x: abs(x[1]['correlation']), 
                                   reverse=True)[:3]
            
            # 生成相关性洞察
            for dim_name, dim_data in top_dimensions:
                corr_value = dim_data['correlation']
                change_pct = dim_data['change_pct']
                
                # 相关性描述
                if corr_value > 0.8:
                    strength = "强"
                elif corr_value > 0.5:
                    strength = "中等"
                else:
                    strength = "弱"
                
                # 相关方向
                if dim_data['pearson_corr'] > 0:
                    direction = "正相关"
                else:
                    direction = "负相关"
                
                # 变化描述
                if change_pct > 30:
                    change_desc = "显著增加"
                elif change_pct > 10:
                    change_desc = "增加"
                elif change_pct < -30:
                    change_desc = "显著减少"
                elif change_pct < -10:
                    change_desc = "减少"
                else:
                    change_desc = "变化不大"
                
                # 格式化列名为用户友好的名称
                friendly_name = dim_name.replace('_', ' ').title()
                
                # 生成洞察
                insight = f"{friendly_name}与销售{direction}({strength}关联)，异常点前该指标{change_desc}({abs(change_pct):.1f}%)，可能是异常原因"
                correlation_insights.append(insight)
        
        # 如果无法找到相关维度，尝试基于时间模式的分析
        if not correlation_insights:
            # 分析时间序列模式
            if time_granularity == 'day':
                # 分析每周模式
                if isinstance(abnormal_date, pd.Timestamp):
                    day_of_week = abnormal_date.dayofweek
                    weekday_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
                    
                    # 获取过去同一周几的数据
                    same_weekdays = []
                    for i in range(current_pos - 1, start_idx - 1, -1):
                        if i >= 0 and isinstance(df_sorted.iloc[i][date_col], pd.Timestamp):
                            if df_sorted.iloc[i][date_col].dayofweek == day_of_week:
                                same_weekdays.append(df_sorted.iloc[i][value_col])
                    
                    if same_weekdays:
                        weekday_avg = sum(same_weekdays) / len(same_weekdays)
                        if weekday_avg > 0:  # 避免除零
                            weekday_change = (abnormal_value - weekday_avg) / weekday_avg * 100
                            
                            if abs(weekday_change) > 20:
                                correlation_insights.append(
                                    f"与过去{len(same_weekdays)}个{weekday_names[day_of_week]}相比，销售变化了{weekday_change:.1f}%，表明周内模式异常"
                                )
                
                # 分析月内模式
                if isinstance(abnormal_date, pd.Timestamp):
                    day_of_month = abnormal_date.day
                    
                    # 分析是否为月初、月中或月末异常
                    if day_of_month <= 5:
                        correlation_insights.append("异常发生在月初(前5天)，可能与月度预算释放或采购周期相关")
                    elif day_of_month >= 25:
                        correlation_insights.append("异常发生在月末(后5天)，可能与月度销售目标冲刺或预算耗尽相关")
            
            elif time_granularity == 'month':
                # 分析季节性模式
                if isinstance(abnormal_date, pd.Timestamp):
                    month = abnormal_date.month
                    season_map = {1: "冬季", 2: "冬季", 3: "春季", 4: "春季", 
                                 5: "春季", 6: "夏季", 7: "夏季", 8: "夏季", 
                                 9: "秋季", 10: "秋季", 11: "秋季", 12: "冬季"}
                    
                    correlation_insights.append(f"异常发生在{month}月({season_map[month]})，可能受季节性因素影响")
        
        # 如果仍未生成任何洞察，提供一个通用分析
        if not correlation_insights:
            correlation_insights.append("未检测到显著的多维度相关性，异常可能由外部因素或未记录变量引起")
    
    except Exception as e:
        print(f"多维度相关性分析出错: {str(e)}")
        correlation_insights.append("多维度相关性分析过程中出现异常，无法提供完整分析")
    
    return correlation_insights

def suggest_anomaly_reasons(df, idx, date_col, value_col, direction, is_consecutive=False, consecutive_score=0, time_granularity='day'):
    """为异常点提供可能的解释"""
    reasons = []
    date = df.iloc[idx][date_col]
    value = df.iloc[idx][value_col]
    
    # 时间粒度映射为中文单位
    time_unit_map = {
        'day': '日',
        'week': '周',
        'month': '月',
        'quarter': '季度',
        'year': '年'
    }
    time_unit = time_unit_map.get(time_granularity, '日')
    
    # 处理连续异常信息
    if is_consecutive:
        if consecutive_score >= 0.9:
            streak_info = f"连续高强度异常 (强度: {consecutive_score:.2f}，所有时间点均异常)"
            reasons.append(streak_info)
            reasons.append(f"可能是持续性系统问题或重大商业活动{time_unit}级影响")
        elif consecutive_score >= 0.7:
            streak_info = f"连续中高强度异常 (强度: {consecutive_score:.2f}，大部分时间点异常)"
            reasons.append(streak_info)
            reasons.append(f"可能是正在发展的{time_unit}度趋势变化或持续性市场波动")
        else:
            streak_info = f"连续异常 (强度: {consecutive_score:.2f}，部分时间点异常)"
            reasons.append(streak_info)
            reasons.append(f"可能是短期{time_unit}度市场波动的开始或局部业务调整")
    
    # 1. 检查是否为节假日
    if isinstance(date, pd.Timestamp):
        is_hol, holiday_name, patterns = is_holiday(date)
        if is_hol:
            if direction == '上升' and patterns:
                reasons.append(f"{holiday_name}: {random.choice(patterns)}")
            elif direction == '下降' and patterns:
                reasons.append(f"{holiday_name}: {random.choice(patterns)}")
            else:
                reasons.append(f"{holiday_name}期间{time_unit}销售{direction}")
            
            # 对连续异常添加节假日相关的特殊解释
            if is_consecutive and consecutive_score > 0.7:
                if direction == '上升':
                    reasons.append(f"节假日期间的持续高{time_unit}销量，可能与假期延长或季节性需求有关")
                else:
                    reasons.append(f"节假日期间的持续低{time_unit}销量，可能与假期延长或供应链中断有关")
    
    # 2. 检查是否为周末 (仅对日粒度有意义)
    if time_granularity == 'day' and isinstance(date, pd.Timestamp) and date.dayofweek >= 5:  # 5=Saturday, 6=Sunday
        if direction == '上升':
            reasons.append("周末购物高峰")
        else:
            reasons.append("周末商业模式变化")
    
    # 3. 检查是否为月初/月末 (仅对日/周粒度有意义)
    if time_granularity in ['day', 'week'] and isinstance(date, pd.Timestamp):
        if date.day <= 3:
            reasons.append("月初消费模式变化")
        elif date.day >= 28:
            reasons.append("月末促销/清仓活动")
    
    # 4. 检查前后变化幅度
    try:
        # 找到前一个和后一个记录的索引
        date_sorted = df.sort_values(by=date_col)
        current_idx = date_sorted.index.get_indexer([df.index[idx]])[0]
        
        if current_idx > 0 and current_idx < len(date_sorted) - 1:
            prev_value = date_sorted.iloc[current_idx-1][value_col]
            next_value = date_sorted.iloc[current_idx+1][value_col]
            
            prev_change = (value - prev_value) / prev_value if prev_value != 0 else 0
            next_change = (next_value - value) / value if value != 0 else 0
            
            # 检查销售模式
            if direction == '上升':
                # 如果是突然的上升异常点，且不是连续上升中
                if prev_change > 0.3 and not (current_idx > 1 and date_sorted.iloc[current_idx-2][value_col] < prev_value):
                    reasons.append(f"突发性{time_unit}销售高峰，可能是促销活动或大客户订单")
                    
                # 检查是否是持续上升中的加速点
                if current_idx > 1 and prev_value > date_sorted.iloc[current_idx-2][value_col]:
                    # 安全检查以避免除以零
                    prev_prev_value = date_sorted.iloc[current_idx-2][value_col]
                    if prev_prev_value != 0:  # 确保不会除以零
                        prev_prev_change = (prev_value - prev_prev_value) / prev_prev_value
                        if prev_change > 2 * prev_prev_change:
                            reasons.append(f"{time_unit}销售增长显著加速，市场反应积极")
                
            elif direction == '下降':
                # 检查这是否是高峰之后的下降
                window_size = min(5, current_idx)
                if window_size >= 3:
                    # 检查前几天是否有明显的高峰
                    window_values = date_sorted.iloc[current_idx-window_size:current_idx][value_col].values
                    peak_idx = np.argmax(window_values)
                    peak_value = window_values[peak_idx]
                    
                    # 如果前几天存在明显高峰，且当前值显著低于高峰
                    if peak_value > value * 1.3 and peak_idx < window_size - 1:
                        days_since_peak = window_size - 1 - peak_idx
                        time_unit_text = {'day': '天', 'week': '周', 'month': '月', 'quarter': '季度', 'year': '年'}.get(time_granularity, '个时间单位')
                        reasons.append(f"临时促销活动后的{time_unit}销量回落（距离峰值约{days_since_peak}{time_unit_text}）")
                
                # 对比同环比下降
                if current_idx >= 7 and isinstance(date, pd.Timestamp) and time_granularity == 'day':
                    try:
                        # 对比上周同一天
                        week_ago_idx = date_sorted[date_sorted[date_col] == date - pd.Timedelta(days=7)].index
                        if len(week_ago_idx) > 0:
                            week_ago_value = date_sorted.loc[week_ago_idx[0], value_col]
                            week_change = (value - week_ago_value) / week_ago_value if week_ago_value != 0 else 0
                            if week_change < -0.3:
                                reasons.append(f"环比上周同期下降{abs(week_change)*100:.1f}%")
                    except:
                        pass
            
            # 以下是之前的检测逻辑，但进行了修改
            # 检查是否是局部峰值后的下降
            if next_change < -0.2:
                # 只有当前不是最高点时，才可能是"销量回落"
                if direction == '下降' and prev_change < 0:
                    # 表示已经开始下降
                    reasons.append(f"下降趋势延续，可能是临时促销效应{time_unit}消退")
            # 检查是否是低谷后的反弹        
            elif prev_change < -0.3 and next_change > 0.2:
                if direction == '上升':
                    reasons.append(f"{time_unit}销售低谷后的反弹，可能是市场调整后恢复")
            
            # 检查是否是连续变化的一部分
            if prev_change > 0.1 and next_change > 0.1:
                if is_consecutive and consecutive_score > 0.7:
                    reasons.append(f"显著的持续{time_unit}增长趋势，可能是产品/市场拓展阶段")
                else:
                    reasons.append(f"{time_unit}销售持续增长趋势")
            elif prev_change < -0.1 and next_change < -0.1:
                if is_consecutive and consecutive_score > 0.7:
                    reasons.append(f"显著的持续{time_unit}下滑趋势，可能面临市场挑战或竞争加剧")
                else:
                    reasons.append(f"{time_unit}销售持续下滑趋势")
    except Exception as e:
        print(f"分析前后变化时出错: {str(e)}")
    
    # 5. 业务导向的异常解析
    business_reasons = analyze_business_impact(df, idx, date_col, value_col, direction, time_granularity)
    # 如果有业务相关的原因解释，添加到原因列表
    if business_reasons:
        # 添加分隔标题
        reasons.append("业务影响分析:")
        # 添加业务原因（最多3个，避免信息过多）
        for br in business_reasons[:3]:
            reasons.append(f"· {br}")
    
    # 6. 多维度数据相关性分析
    multidim_insights = analyze_multidimensional_correlation(df, date_col, value_col, idx, time_granularity)
    if multidim_insights:
        # 添加分隔标题
        reasons.append("多维度相关性分析:")
        # 添加相关性洞察（最多3个）
        for insight in multidim_insights[:3]:
            reasons.append(f"· {insight}")

    # 7. 如果没有找到具体原因，提供常见解释
    if len(reasons) <= (2 if is_consecutive else 0):
        if direction == '上升':
            common_up_reasons = [f"可能的{time_unit}促销活动", f"{time_unit}市场需求突增", f"特殊事件带动{time_unit}销售"]
            reasons.append(random.choice(common_up_reasons))
        else:
            common_down_reasons = [f"{time_unit}供应链可能中断", f"{time_unit}市场竞争加剧", f"{time_unit}消费者行为变化"]
            reasons.append(random.choice(common_down_reasons))
    
    return reasons

def detect_anomalies(df, date_col, value_col, z_threshold=2.5, detect_consecutive=True, time_granularity='day'):
    """
    检测销售数据中的异常点，使用多维度异常评分，包括连续异常检测
    
    参数:
    - df: 数据框
    - date_col: 日期列名
    - value_col: 值列名
    - z_threshold: 异常阈值，默认2.5
    - detect_consecutive: 是否检测连续异常，默认True
    - time_granularity: 时间粒度，可选值: 'day', 'week', 'month', 'quarter', 'year'
    """
    # 检查数据量，大数据集进行抽样
    if len(df) > 10000:
        # 对于大数据集，仅取周期性样本
        sampling_rate = max(1, len(df) // 5000)  # 最多5000个样本点
        df_sample = df.iloc[::sampling_rate].copy()
        print(f"异常检测: 数据量大，从{len(df)}行抽样到{len(df_sample)}行")
    else:
        df_sample = df.copy()
    
    # 确保日期列是时间类型 - 对业务分析很重要
    try:
        if not pd.api.types.is_datetime64_any_dtype(df_sample[date_col]):
            df_sample[date_col] = pd.to_datetime(df_sample[date_col], errors='coerce')
    except Exception as e:
        print(f"转换日期列时出错: {str(e)}")
        
    # 使用多维度异常评分
    combined_scores, individual_scores, threshold_votes, anomaly_directions = calculate_multidimensional_anomaly_score(
        df_sample[value_col]
    )
    
    #print("\n===== 异常检测均衡性分析 =====")
    # 统计上升和下降的异常
    up_anomalies = np.sum(anomaly_directions > 0)
    down_anomalies = np.sum(anomaly_directions < 0)
    total_potential = np.sum((combined_scores > z_threshold) & (threshold_votes >= 2))
    #print(f"检测到的潜在异常总数: {total_potential}")
    
    # 安全计算百分比，避免除以零
    up_percent = (up_anomalies/total_potential*100) if total_potential > 0 else 0
    down_percent = (down_anomalies/total_potential*100) if total_potential > 0 else 0
    
    #print(f"上升异常: {up_anomalies} ({up_percent:.1f}%)")
    #print(f"下降异常: {down_anomalies} ({down_percent:.1f}%)")
    
    if total_potential > 0 and up_anomalies > down_anomalies * 2 and down_anomalies > 0:
        print("警告: 检测到的上升异常显著多于下降异常，可能存在偏差")
        # 调整下降异常的敏感度
        z_threshold_down = z_threshold * 0.85  # 为下降异常提供更宽松的阈值
        print(f"正在调整下降异常阈值为: {z_threshold_down:.2f} (标准阈值: {z_threshold})")
    elif down_anomalies > up_anomalies * 2 and up_anomalies > 0:
        print("警告: 检测到的下降异常显著多于上升异常，可能存在偏差")
        # 调整上升异常的敏感度
        z_threshold_up = z_threshold * 0.85  # 为上升异常提供更宽松的阈值
        print(f"正在调整上升异常阈值为: {z_threshold_up:.2f} (标准阈值: {z_threshold})")
    else:
        z_threshold_up = z_threshold_down = z_threshold
        #print("检测平衡正常，上升和下降异常分布合理")
    
    # 找出超过阈值的基础异常点 - 区分上升和下降
    anomaly_markers = np.zeros(len(df_sample), dtype=bool)
    for i in range(len(df_sample)):
        if anomaly_directions[i] > 0:  # 上升异常
            anomaly_markers[i] = (combined_scores[i] > z_threshold_up) and (threshold_votes[i] >= 2)
        elif anomaly_directions[i] < 0:  # 下降异常
            anomaly_markers[i] = (combined_scores[i] > z_threshold_down) and (threshold_votes[i] >= 2)
    
    # 新增: 检测尖峰模式异常 (上升后立即下降的模式)
    #print("\n===== 尖峰模式检测 =====")
    #print("检测销售数据中的尖峰模式(突然上升后迅速下降)")
    
    # 初始化尖峰模式标记
    pattern_markers = {
        "spike_pattern": np.zeros(len(df_sample), dtype=bool),  # 是否为尖峰模式
        "spike_prominence": np.zeros(len(df_sample))            # 尖峰显著程度
    }
    
    # 确保数据按日期排序
    df_sorted = df_sample.sort_values(by=date_col).reset_index(drop=True)
    
    # 提取销售值并处理无效数据
    values = df_sorted[value_col].values
    dates = df_sorted[date_col].values
    
    # 检查无效数据
    valid_mask = ~np.isnan(values) & (values != 0)  # 非NaN且非零值
    #print(f"检测到{np.sum(~valid_mask)}个无效数据点(NaN或零值)，将在尖峰检测中排除")
    
    # 只在有足够有效数据点的情况下进行尖峰检测
    if np.sum(valid_mask) > 5:  # 至少需要5个有效点
        # 创建有效数据的副本以进行计算
        valid_indices = np.where(valid_mask)[0]
        valid_values = values[valid_mask]
        valid_dates = dates[valid_mask]
        
        # 计算移动平均线(作为基线)，仅使用有效数据
        window_size = max(3, min(7, len(valid_values) // 20))  # 窗口大小根据有效数据量调整
        if len(valid_values) > window_size * 2:
            # 计算移动平均作为基线
            baseline = np.convolve(valid_values, np.ones(window_size)/window_size, mode='same')
            
            # 查找潜在尖峰（仅在有效数据中查找）
            spike_count = 0
            for i in range(window_size, len(valid_values) - window_size):
                # 确保前后点都是有效的连续数据
                if i > 0 and i < len(valid_values) - 1:
                    # 获取真实索引
                    real_idx = valid_indices[i]
                    # 获取前后的真实索引
                    prev_real_idx = valid_indices[i-1] if i-1 >= 0 else None
                    next_real_idx = valid_indices[i+1] if i+1 < len(valid_indices) else None
                    
                    # 计算前后数据点的间隔（以天为单位）
                    date_gaps_ok = True
                    if prev_real_idx is not None and next_real_idx is not None:
                        try:
                            # 将日期转换为datetime对象
                            current_date = pd.to_datetime(dates[real_idx])
                            prev_date = pd.to_datetime(dates[prev_real_idx])
                            next_date = pd.to_datetime(dates[next_real_idx])
                            
                            # 计算日期间隔（天数）
                            days_to_prev = (current_date - prev_date).days
                            days_to_next = (next_date - current_date).days
                            
                            # 确保日期间隔合理（防止长时间无数据的情况）
                            max_gap = 7  # 最大允许间隔7天
                            if days_to_prev > max_gap or days_to_next > max_gap:
                                date_gaps_ok = False
                        except:
                            # 日期转换失败，保守处理
                            date_gaps_ok = False
                    
                    # 只有当日期间隔合理时才进行尖峰检测
                    if date_gaps_ok and valid_values[i] > valid_values[i-1] and valid_values[i] > valid_values[i+1]:
                        # 计算相对于移动平均的显著程度
                        prominence = (valid_values[i] - baseline[i]) / (baseline[i] if baseline[i] > 0 else 1)
                        
                        # 计算相对于前后点的显著程度
                        rel_to_prev = (valid_values[i] - valid_values[i-1]) / (valid_values[i-1] if valid_values[i-1] > 0 else 1)
                        rel_to_next = (valid_values[i] - valid_values[i+1]) / (valid_values[i+1] if valid_values[i+1] > 0 else 1)
                        
                        # 综合显著程度
                        overall_prominence = (prominence + rel_to_prev + rel_to_next) / 3
                        
                        # 如果尖峰足够显著，标记为尖峰模式
                        if overall_prominence > 0.3:  # 尖峰显著度阈值
                            # 找到排序前的索引（在原始df_sample中的位置）
                            orig_idx = df_sorted.index[real_idx]
                            # 标记为尖峰模式
                            pattern_markers["spike_pattern"][orig_idx] = True
                            pattern_markers["spike_prominence"][orig_idx] = overall_prominence
                            spike_count += 1
            
            #print(f"检测到{spike_count}个尖峰模式")
            
            # 尖峰显著度分布
            if spike_count > 0:
                high_prom = np.sum(pattern_markers["spike_prominence"] > 0.6)
                med_prom = np.sum((pattern_markers["spike_prominence"] > 0.4) & (pattern_markers["spike_prominence"] <= 0.6))
                low_prom = np.sum((pattern_markers["spike_prominence"] > 0.3) & (pattern_markers["spike_prominence"] <= 0.4))
                
                #print(f"尖峰显著度分布:")
                #print(f"- 高显著度(>0.6): {high_prom}个")
                #print(f"- 中显著度(0.4-0.6): {med_prom}个")
                #print(f"- 低显著度(0.3-0.4): {low_prom}个")
        else:
            print(f"有效数据点不足({np.sum(valid_mask)}个)，无法可靠计算尖峰模式")
    else:
        print(f"有效数据点不足({np.sum(valid_mask)}个)，跳过尖峰模式检测")
    
    # 将尖峰模式添加到异常标记中
    combined_anomaly_markers = anomaly_markers.copy()
    for i in range(len(df_sample)):
        if pattern_markers["spike_pattern"][i]:
            combined_anomaly_markers[i] = True
    
    # 执行连续异常检测（如果启用）
    consecutive_scores = np.zeros(len(df_sample))
    is_in_streak = np.zeros(len(df_sample), dtype=bool)
    
    if detect_consecutive and len(df_sample) >= 3:
        #print("\n===== 连续异常检测 =====")
        #print("检测连续出现的异常模式，这可能表示系统性问题而非随机波动")
        
        # 确保日期列是时间戳格式
        try:
            if not pd.api.types.is_datetime64_any_dtype(df_sample[date_col]):
                df_sample[date_col] = pd.to_datetime(df_sample[date_col], errors='coerce')
            
            # 排序数据 
            df_sample = df_sample.sort_values(by=date_col).reset_index(drop=False)
            
            # 获取原始索引的映射
            original_index = df_sample['index'].values
            df_sample = df_sample.drop(columns=['index'])
            
            # 检测连续异常 - 使用合并后的异常标记
            consecutive_scores, is_in_streak = detect_consecutive_anomalies(
                df_sample, 
                date_col, 
                combined_anomaly_markers,
                window_size=3,  # 考虑3天窗口
                min_anomalies=2  # 要求至少2天异常
            )
            
            # 统计连续异常数量
            streak_count = np.sum(is_in_streak)
            if streak_count > 0:
                #print(f"\n共检测到{streak_count}个连续异常点:")
                # 计算平均强度
                avg_strength = np.mean(consecutive_scores[is_in_streak])
                #print(f"连续异常平均强度: {avg_strength:.2f}")
                
                # 解释强度含义
                if avg_strength > 0.9:
                    #print("整体呈现高度系统性异常模式，建议重点关注")
                    pass
                elif avg_strength > 0.7:
                    #print("存在较显著的系统性异常模式")
                    pass
                else:
                    #print("存在轻微的系统性异常模式")
                    pass
                    
                #print("\n连续异常强度理解指南:")
                #print("- 1.00: 所有时间点都异常，极有可能是系统性问题")
                #print("- 0.67: 3天中有2天异常，可能是短期波动或开始的趋势变化")
                #print("- 多个连续区间：表示周期性异常，可能与业务周期相关")
            else:
                #print("未检测到连续异常模式，所有异常点均为独立事件")
                pass
                
        except Exception as e:
            print(f"连续异常检测时出错: {str(e)}")
    
    # 找出异常点 - 使用合并后的异常标记
    if detect_consecutive:
        anomalies_idx = np.where(combined_anomaly_markers | (consecutive_scores > 0.5))[0]
    else:
        anomalies_idx = np.where(combined_anomaly_markers)[0]
        
    anomalies = []
    
    # 业务分析时需要额外信息 - 启用更多分析
    #print("\n===== 业务导向的异常分析 =====")
    #print("分析异常对业务的具体影响，包括销售路径、转化率和投资回报评估")
    
    # 统计上升、下降和尖峰异常点
    final_up = 0
    final_down = 0
    final_spike = 0  # 新增：尖峰异常统计
    
    # 限制异常点数量，避免过多
    if len(anomalies_idx) > 20:
        #print(f"\n检测到{len(anomalies_idx)}个异常点，将只显示最显著的20个")
        # 创建组合分数，包含连续异常加权和尖峰模式优先级
        final_scores = combined_scores.copy()
        if detect_consecutive:
            final_scores += consecutive_scores * 1.5  # 连续异常权重更高
        
        # 提升尖峰模式的优先级
        for i in range(len(df_sample)):
            if pattern_markers["spike_pattern"][i]:
                # 尖峰显著度越高，优先级越高
                final_scores[i] += pattern_markers["spike_prominence"][i] * 2.0
            
        # 分别为不同类型异常排序
        up_idx = [idx for idx in anomalies_idx if anomaly_directions[idx] > 0 and not pattern_markers["spike_pattern"][idx]]
        down_idx = [idx for idx in anomalies_idx if anomaly_directions[idx] < 0 and not pattern_markers["spike_pattern"][idx]]
        spike_idx = [idx for idx in anomalies_idx if pattern_markers["spike_pattern"][idx]]
        
        # 按显著程度排序
        if len(up_idx) > 0:
            up_sorted = up_idx[np.argsort(final_scores[up_idx])[::-1]]
        else:
            up_sorted = []
            
        if len(down_idx) > 0:
            down_sorted = down_idx[np.argsort(final_scores[down_idx])[::-1]]
        else:
            down_sorted = []
            
        if len(spike_idx) > 0:
            spike_sorted = spike_idx[np.argsort(pattern_markers["spike_prominence"][spike_idx])[::-1]]
        else:
            spike_sorted = []
        
        # 平衡选择不同类型的异常
        total_slots = min(20, len(anomalies_idx))
        
        # 计算各类型异常的占比
        total_anomalies = len(up_idx) + len(down_idx) + len(spike_idx)
        
        # 计算各类型应分配的名额
        if total_anomalies > 0:
            # 确保尖峰异常优先展示
            spike_slots = min(len(spike_idx), max(1, int(total_slots * 0.3)))  # 至少1个，最多30%
            
            # 剩余名额按比例分配给上升和下降异常
            remaining_slots = total_slots - spike_slots
            up_prop = len(up_idx) / max(1, (len(up_idx) + len(down_idx)))
            
            if len(up_idx) > 0 and len(down_idx) > 0:
                up_slots = min(len(up_idx), max(1, int(remaining_slots * up_prop)))
                down_slots = min(len(down_idx), remaining_slots - up_slots)
            elif len(up_idx) > 0:
                up_slots = min(len(up_idx), remaining_slots)
                down_slots = 0
            elif len(down_idx) > 0:
                up_slots = 0
                down_slots = min(len(down_idx), remaining_slots)
            else:
                up_slots = down_slots = 0
            
            #print(f"平衡选择: {up_slots}个上升异常，{down_slots}个下降异常，{spike_slots}个尖峰模式异常")
            
            # 选择最终要展示的异常点
            selected_up = up_sorted[:up_slots] if up_slots > 0 else []
            selected_down = down_sorted[:down_slots] if down_slots > 0 else []
            selected_spike = spike_sorted[:spike_slots] if spike_slots > 0 else []
            
            # 合并选定的点
            anomalies_idx = np.concatenate((selected_up, selected_down, selected_spike)) if len(selected_up) + len(selected_down) + len(selected_spike) > 0 else anomalies_idx
            
            # 记录最终各类型异常数量
            final_up = len(selected_up)
            final_down = len(selected_down)
            final_spike = len(selected_spike)
        else:
            # 如果按类型分类后没有异常，则使用原始排序
            sorted_idx = np.argsort(final_scores[anomalies_idx])[::-1][:20]
            anomalies_idx = anomalies_idx[sorted_idx]
    else:
        # 计算最终选择的不同类型异常数量
        for idx in anomalies_idx:
            if pattern_markers["spike_pattern"][idx]:
                final_spike += 1
            elif anomaly_directions[idx] > 0:
                final_up += 1
            elif anomaly_directions[idx] < 0:
                final_down += 1
    
    # 更新输出信息，添加尖峰异常数量
    #print(f"最终选择的异常: {final_up}个上升，{final_down}个下降，{final_spike}个尖峰模式")
    
    # 格式化日期的函数，根据时间粒度显示不同的格式
    def format_date_by_granularity(date_value, granularity):
        if not isinstance(date_value, pd.Timestamp):
            try:
                date_value = pd.to_datetime(date_value)
            except:
                return str(date_value)  # 无法转换时返回原始字符串
        
        # 根据不同的时间粒度格式化日期
        if granularity == 'day':
            return date_value.strftime('%Y-%m-%d')
        elif granularity == 'week':
            # 获取周数 (ISO周)，减去1以符合业务习惯
            iso_year, iso_week, iso_day = date_value.isocalendar()
            week_num = iso_week - 1
            
            # 计算上一周的日期范围（周一至周日）
            # 首先计算ISO周的周一日期
            iso_monday = date_value - pd.Timedelta(days=iso_day-1)
            
            # 将日期范围提前一周，然后整体后移一天
            prev_week_monday = iso_monday - pd.Timedelta(days=6)  # 减少了一天
            prev_week_sunday = iso_monday  # 现在是本周一，而不是上周日
            
            # 根据减1后的周数调整日期和年份
            if week_num == 0:
                # 上一年的最后一周
                prev_year = date_value.year - 1
                # 上一年的最后一天
                last_day_prev_year = pd.Timestamp(f"{prev_year}-12-31")
                last_iso_year, last_iso_week, last_iso_day = last_day_prev_year.isocalendar()
                week_num = last_iso_week
                year = prev_year
            else:
                year = iso_year
                
            # 跨年处理：如果上周的一部分在上一年
            if prev_week_monday.year < year:
                # 使用上周一的年份作为显示年份
                year = prev_week_monday.year
            
            # 格式化日期范围
            date_range = f"{prev_week_monday.month}月{prev_week_monday.day}日-{prev_week_sunday.month}月{prev_week_sunday.day}日"
            
            return f"{year}年{week_num}周（{date_range}）"
        elif granularity == 'month':
            return date_value.strftime('%Y年%m月')
        elif granularity == 'quarter':
            # 计算季度
            quarter = (date_value.month - 1) // 3 + 1
            return f"{date_value.year}年Q{quarter}"
        elif granularity == 'year':
            return f"{date_value.year}年"
        else:
            return date_value.strftime('%Y-%m-%d')  # 默认使用日期格式
    
    for idx in anomalies_idx:
        anomaly_date = df_sample.iloc[idx][date_col]
        value = df_sample.iloc[idx][value_col]
        
        # 获取各维度分数
        zscore = individual_scores["zscore"][idx]
        iqr_score = individual_scores["iqr"][idx]
        mad_score = individual_scores["mad"][idx]
        
        # 判断异常类型和方向
        if pattern_markers["spike_pattern"][idx]:
            # 尖峰模式异常
            anomaly_type = "尖峰模式"
            direction = "尖峰"  # 特殊方向标记
        else:
            # 常规上升/下降异常
            anomaly_type = "常规异常"
            direction = '上升' if anomaly_directions[idx] > 0 else '下降'
        
        # 尝试找出可能的原因，传递时间粒度参数
        if anomaly_type == "尖峰模式":
            # 尖峰模式特殊处理：分析前后数据点找出尖峰原因
            prominence = pattern_markers["spike_prominence"][idx]
            reasons = [
                f"检测到销售尖峰模式(显著度: {prominence:.2f})",
                f"突然上升后快速下降的模式，典型的促销活动或特殊事件特征"
            ]
            
            # 根据显著程度提供不同级别的解释
            if prominence > 0.6:
                reasons.append(f"非常显著的尖峰，可能是重大促销活动或特殊事件影响")
            elif prominence > 0.4:
                reasons.append(f"中等显著的尖峰，可能是常规促销或季节性事件")
            else:
                reasons.append(f"轻微尖峰，可能是小规模促销或临时因素")
                
            # 添加业务影响分析
            try:
                # 尖峰后续趋势分析
                df_sorted = df_sample.sort_values(by=date_col).reset_index(drop=True)
                current_pos = df_sorted[df_sorted[date_col] == anomaly_date].index[0]
                
                # 获取尖峰点前后的数据进行分析
                pre_spike_idx = max(0, current_pos - 2)
                post_spike_idx = min(len(df_sorted) - 1, current_pos + 2)
                
                if current_pos > 0 and current_pos < len(df_sorted) - 1:
                    pre_value = df_sorted.iloc[pre_spike_idx][value_col]
                    post_value = df_sorted.iloc[post_spike_idx][value_col]
                    
                    if post_value < pre_value:
                        reasons.append(f"尖峰后销售下降至低于尖峰前水平，可能是透支了未来需求")
                    elif post_value > pre_value:
                        reasons.append(f"尖峰后销售维持在高于尖峰前水平，表明促销带来了持续增长效应")
                    else:
                        reasons.append(f"尖峰后销售回归至基准水平，属于典型的短期促销模式")
            except Exception as e:
                print(f"尖峰后续分析出错: {str(e)}")
        else:
            # 常规异常处理
            reasons = suggest_anomaly_reasons(df_sample, idx, date_col, value_col, direction, 
                                          is_consecutive=is_in_streak[idx], 
                                          consecutive_score=consecutive_scores[idx],
                                          time_granularity=time_granularity)
        
        # 使用根据时间粒度格式化的日期
        formatted_date = format_date_by_granularity(anomaly_date, time_granularity)
        
        # 添加异常评分信息
        anomaly_info = {
            'date': formatted_date,
            'raw_date': anomaly_date.strftime('%Y-%m-%d') if isinstance(anomaly_date, pd.Timestamp) else str(anomaly_date),
            'value': float(value),
            'anomaly_score': float(combined_scores[idx]),
            'zscore': float(zscore),
            'iqr_score': float(iqr_score),
            'mad_score': float(mad_score),
            'votes': int(threshold_votes[idx]),
            'direction': direction,
            'anomaly_type': anomaly_type,  # 新增：异常类型
            'consecutive': float(consecutive_scores[idx]) if detect_consecutive else 0,
            'is_in_streak': bool(is_in_streak[idx]) if detect_consecutive else False,
            'spike_prominence': float(pattern_markers["spike_prominence"][idx]) if pattern_markers["spike_pattern"][idx] else 0.0,  # 新增：尖峰显著度
            'reasons': reasons
        }
        
        anomalies.append(anomaly_info)
    
    return anomalies

def get_analysis_suggestions():
    """根据上传的数据文件，提供分析建议"""
    try:
        # 重置取消标志 - 每次新的请求都重置
        global analysis_cancelled
        analysis_cancelled = False
        
        # 创建唯一的任务ID
        task_id = str(uuid.uuid4())
        
        # 获取用户信息
        username = session.get('username', '匿名用户')
        client_ip = request.remote_addr
        
        # 记录任务
        ongoing_analyses[task_id] = {
            'cancelled': False,
            'start_time': datetime.now(),
            'username': username,
            'ip': client_ip,
            'operation': 'file_preprocessing'
        }
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '请上传数据文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'})
        
        # 获取用户选择的工作表（如果有）
        sheet_name = request.form.get('sheet_name', None)
        #print(f"用户选择的工作表: {sheet_name}")
        
        # 保存文件
        temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 
                              f"temp_{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}")
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        # 检查是否取消
        if is_analysis_cancelled(task_id):
            print("文件保存后检测到取消请求，中止处理")
            return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
        
        # 首先检查Excel文件的工作表
        sheet_names = []
        try:
            # 使用pandas获取所有工作表名称
            xl = pd.ExcelFile(temp_path)
            sheet_names = xl.sheet_names
            
            # 检查是否取消
            if is_analysis_cancelled(task_id):
                print("工作表检查后检测到取消请求，中止处理")
                return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
            
            # 如果指定了工作表并且存在，则使用它
            if sheet_name and sheet_name in sheet_names:
                selected_sheet = sheet_name
                #print(f"使用用户指定的工作表: {selected_sheet}")
            # 否则如果有工作表，默认使用第一个
            elif sheet_names:
                selected_sheet = sheet_names[0]
                #print(f"使用默认工作表: {selected_sheet}")
            else:
                selected_sheet = 0
                #print("没有检测到工作表，使用默认索引0")
                
            multiple_sheets = len(sheet_names) > 1
            
            #print(f"检测到工作表: {sheet_names}, 选择的工作表: {selected_sheet}, 多工作表: {multiple_sheets}")
        except Exception as e:
            print(f"检查工作表时出错: {str(e)}")
            selected_sheet = 0 if sheet_name is None else sheet_name
            multiple_sheets = False
        
        # 再次检查是否取消
        if is_analysis_cancelled(task_id):
            print("工作表确认后检测到取消请求，中止处理")
            return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
        
        try:
            # 读取Excel文件，先尝试识别数据量级
            try:
                # 只读取头部获取列信息
                df_sample = pd.read_excel(temp_path, nrows=5, sheet_name=selected_sheet)
                
                # 检查是否取消
                if is_analysis_cancelled(task_id):
                    print("样本读取后检测到取消请求，中止处理")
                    return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
                
                # 估算数据行数 (仅适用于单表单Excel，多表单会有所不同)
                try:
                    # 尝试使用xlrd（仅支持旧版.xls格式）
                    workbook = xlrd.open_workbook(temp_path, on_demand=True)
                    if isinstance(selected_sheet, str):
                        try:
                            sheet = workbook.sheet_by_name(selected_sheet)
                        except:
                            # 如果找不到指定名称的工作表，使用第一个
                            sheet = workbook.sheet_by_index(0)
                    else:
                        sheet = workbook.sheet_by_index(0)
                    estimated_rows = sheet.nrows - 1  # 减去标题行
                    workbook.release_resources()
                    #print(f"使用xlrd估算行数: {estimated_rows}")
                except Exception as xlrd_error:
                    # 检查是否取消
                    if is_analysis_cancelled(task_id):
                        print("行数估算过程中检测到取消请求，中止处理")
                        return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
                    
                    # xlrd失败，尝试使用openpyxl（支持新版.xlsx格式）
                    try:
                        workbook = openpyxl.load_workbook(temp_path, read_only=True)
                        if isinstance(selected_sheet, str) and selected_sheet in workbook.sheetnames:
                            sheet = workbook[selected_sheet]
                        else:
                            # 使用活动工作表
                            sheet = workbook.active
                        # openpyxl在read_only模式下无法直接获取行数，使用最大行作为估计
                        estimated_rows = sheet.max_row - 1  # 减去标题行
                        workbook.close()
                        #print(f"使用openpyxl估算行数: {estimated_rows}")
                    except Exception as openpyxl_error:
                        # 检查是否取消
                        if is_analysis_cancelled(task_id):
                            print("行数估算过程中检测到取消请求，中止处理")
                            return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
                        
                        # 如果两种方法都失败，通过pandas读取行数
                        print(f"估算行数失败，使用pandas读取全部数据: {str(xlrd_error)}")
                        df = pd.read_excel(temp_path, sheet_name=selected_sheet)
                        estimated_rows = len(df)
                        #print(f"通过pandas获取行数: {estimated_rows}")
                        # 已经读取了完整数据，标记已读取
                        df_already_loaded = True
                    
                #print(f"估算行数: {estimated_rows}")
                
                # 检查是否取消
                if is_analysis_cancelled(task_id):
                    print("行数估算后检测到取消请求，中止处理")
                    return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
                
                # 数据量超大时进行抽样
                sample_size = 10000  # 抽样大小
                if not locals().get('df_already_loaded', False):  # 检查是否已读取
                    if estimated_rows > 50000:
                        #print(f"检测到大数据集: 约 {estimated_rows} 行，将进行抽样分析")
                        df = pd.read_excel(temp_path, sheet_name=selected_sheet, 
                                          skiprows=lambda x: x > 0 and x % (estimated_rows // sample_size + 1) != 0)
                    else:
                        df = pd.read_excel(temp_path, sheet_name=selected_sheet)
                    
            except Exception as e:
                # 检查是否取消
                if is_analysis_cancelled(task_id):
                    print("行数估算失败处理中检测到取消请求，中止处理")
                    return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
                
                print(f"估算行数时出错: {str(e)}")
                # 如果无法估算，直接读取全部数据
                df = pd.read_excel(temp_path, sheet_name=selected_sheet)
                estimated_rows = len(df)
            
            # 检查是否取消
            if is_analysis_cancelled(task_id):
                print("数据加载后检测到取消请求，中止处理")
                return jsonify({'success': False, 'message': '操作已取消', 'cancelled': True})
            
            # 在整个函数内抑制警告
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                
                # 检测可能的日期列
                date_columns = []
                
                # 数据量大时优化日期列检测
                row_count = len(df)
                #print(f"实际读取行数: {row_count}")
                
                # 更严格的日期列检测
                for col in df.columns:
                    # 首先检查列名是否包含日期相关关键词
                    col_lower = str(col).lower()
                    date_related_keywords = ['date', 'time', 'day', 'month', 'year', '日期', '时间', '日', '月', '年']
                    
                    # 仅对列名中包含日期关键词的列进行进一步检测
                    if any(keyword in col_lower for keyword in date_related_keywords):
                        # 取样本进行检测
                        sample_vals = df[col].dropna().head(100)  # 取前100个非空值
                        
                        if len(sample_vals) > 0:
                            # 尝试转换为日期
                            try:
                                converted_dates = pd.to_datetime(sample_vals, errors='coerce')
                                # 检查转换成功率，至少80%能成功转换才认为是日期列
                                success_rate = converted_dates.notna().mean()
                                if success_rate >= 0.8:
                                    #print(f"列 '{col}' 被识别为日期列，转换成功率: {success_rate:.2f}")
                                    date_columns.append(col)
                                else:
                                    print(f"列 '{col}' 虽然名称含日期关键词，但无法转换为日期（成功率: {success_rate:.2f}）")
                            except Exception as e:
                                print(f"尝试将列 '{col}' 转换为日期时出错: {str(e)}")
                    
                    # 特殊处理：对于列名中不包含日期关键词但可能是日期的列
                    # 常见的日期列名: order_date, created_at, 下单时间, 支付时间等
                    elif any(keyword in col_lower for keyword in ['order', 'created', 'updated', 'paid', '下单', '支付', '发货', '签收']):
                        # 取样本进行检测
                        sample_vals = df[col].dropna().head(100)
                        
                        if len(sample_vals) > 0:
                            # 检查样本中是否包含日期样式的字符串
                            sample_strings = sample_vals.astype(str)
                            
                            # 检查是否包含常见的日期分隔符 (/, -, :)
                            date_pattern = r'\d{1,4}[-/年]\d{1,2}[-/月]\d{1,2}'
                            date_time_pattern = r'\d{1,4}[-/年]\d{1,2}[-/月]\d{1,2}.{0,1}\d{1,2}[:时]\d{1,2}'
                            
                            # 随机抽取10个样本检查
                            if len(sample_strings) > 10:
                                check_samples = sample_strings.sample(10)
                            else:
                                check_samples = sample_strings
                                
                            date_matches = [bool(re.search(date_pattern, s) or re.search(date_time_pattern, s)) for s in check_samples]
                            date_match_rate = sum(date_matches) / len(date_matches)
                            
                            if date_match_rate > 0.5:  # 如果超过50%符合日期模式
                                # 尝试转换
                                try:
                                    converted_dates = pd.to_datetime(sample_vals, errors='coerce')
                                    success_rate = converted_dates.notna().mean()
                                    if success_rate >= 0.8:
                                        #print(f"列 '{col}' 虽然名称不含日期关键词，但检测为日期列，转换成功率: {success_rate:.2f}")
                                        date_columns.append(col)
                                except:
                                    pass
                
                #print(f"检测到的日期列: {date_columns}")
                
                # 检测可能的数值列
                value_columns = []
                
                # 遍历所有列查找数值列（包括文本格式的数字）
                for col in df.columns:
                    if col in date_columns:
                        continue  # 跳过日期列
                        
                    # 检查是否是原生数值类型
                    is_numeric = pd.api.types.is_numeric_dtype(df[col])
                    
                    # 如果不是原生数值类型，尝试转换
                    try_conversion = False
                    if not is_numeric:
                        # 尝试先转换为数值类型
                        try:
                            # 抽取样本进行测试转换
                            sample = df[col].dropna().head(100)
                            if len(sample) > 0:
                                # 检查样本中是否大部分可以转换为数字
                                # 先把明显的千分位分隔符去掉
                                numeric_sample = sample.astype(str).str.replace(',', '')
                                # 去掉可能的货币符号和空格
                                numeric_sample = numeric_sample.str.replace('[$¥€£]', '', regex=True).str.strip()
                                
                                # 检查是否包含数字模式
                                numeric_pattern = r'^-?\d+(\.\d+)?$'  # 匹配整数或小数
                                currency_pattern = r'^-?[$¥€£]?\s*\d+(,\d{3})*(\.\d+)?$'  # 匹配货币格式
                                
                                # 检查每个值是否匹配数字模式
                                pattern_matches = [bool(re.match(numeric_pattern, s) or re.match(currency_pattern, s)) 
                                                 for s in numeric_sample]
                                pattern_match_rate = sum(pattern_matches) / len(pattern_matches) if pattern_matches else 0
                                
                                # 尝试转换为数值
                                converted = pd.to_numeric(numeric_sample, errors='coerce')
                                non_na_ratio = converted.notna().mean()
                                
                                # 从两个指标来判断是否为数值列
                                # 1. 正则表达式模式匹配率
                                # 2. 转换为数值的成功率
                                is_likely_numeric = pattern_match_rate >= 0.7 or non_na_ratio >= 0.7
                                
                                # 如果大部分值可以转换为数值，则认为这是一个数值列
                                if is_likely_numeric:
                                    #print(f"列 '{col}' 被检测为文本格式的数值列，模式匹配率: {pattern_match_rate:.2f}, 转换成功率: {non_na_ratio:.2f}")
                                    try_conversion = True
                                    is_numeric = True
                        except Exception as e:
                            print(f"尝试将列 '{col}' 转换为数值类型时出错: {str(e)}")
                            
                    # 特殊处理: 列名含有明显数值关键词但未被识别为数值的列
                    if not is_numeric:
                        value_keywords = ['price', 'amount', 'quantity', 'count', 'sum', 'total', 'cost', 'fee', 
                                         '价', '金额', '数量', '成本', '费用', '合计', '总计', '件数', '单价']
                        
                        if any(keyword in str(col).lower() for keyword in value_keywords):
                            # 再次尝试更宽松的转换
                            try:
                                sample = df[col].dropna().head(100)
                                if len(sample) > 0:
                                    # 更宽松的处理方式
                                    numeric_sample = sample.astype(str).str.replace(r'[^\d.-]', '', regex=True)
                                    # 去掉空字符串
                                    numeric_sample = numeric_sample[numeric_sample != '']
                                    if len(numeric_sample) > 0:
                                        # 尝试转换
                                        converted = pd.to_numeric(numeric_sample, errors='coerce')
                                        non_na_ratio = converted.notna().mean()
                                        
                                        if non_na_ratio >= 0.6:  # 更宽松的阈值
                                            #print(f"列 '{col}' 因名称含数值关键词被检测为数值列，转换成功率: {non_na_ratio:.2f}")
                                            try_conversion = True
                                            is_numeric = True
                            except Exception as e:
                                print(f"针对数值关键词列 '{col}' 的特殊处理失败: {str(e)}")
                    
                    # 如果是数值类型或可以转换为数值，进行进一步处理
                    if is_numeric:
                        try:
                            col_data = df[col]
                            
                            # 如果需要转换文本为数值
                            if try_conversion:
                                # 创建转换函数
                                def convert_to_numeric(val):
                                    if pd.isna(val):
                                        return np.nan
                                    try:
                                        # 如果已经是数值类型，直接返回
                                        if isinstance(val, (int, float)):
                                            return val
                                            
                                        # 处理字符串类型
                                        val_str = str(val)
                                        
                                        # 移除货币符号、千分位分隔符等
                                        val_str = val_str.replace(',', '')  # 去掉千分位分隔符
                                        val_str = re.sub(r'[$¥€£]', '', val_str)  # 去掉货币符号
                                        val_str = val_str.strip()  # 去掉前后空格
                                        
                                        # 处理百分比
                                        if '%' in val_str:
                                            val_str = val_str.replace('%', '')
                                            return float(val_str) / 100
                                            
                                        # 转换为浮点数
                                        return float(val_str)
                                    except:
                                        return np.nan
                                
                                # 应用转换函数
                                col_data = df[col].apply(convert_to_numeric)
                            
                            # 确保转换后的数据能够计算统计值
                            col_data = pd.to_numeric(col_data, errors='coerce')
                            
                            # 计算非零值比例
                            non_zero_count = (col_data != 0).sum()
                            total_count = col_data.count()  # 只计算非NA值
                            non_zero_ratio = non_zero_count / total_count if total_count > 0 else 0
                            
                            # 放宽非零值比例的限制
                            if non_zero_ratio > 0.05:  # 只要有5%的非零值就接受
                                # 计算统计值
                                col_mean = col_data.mean()
                                col_min = col_data.min()
                                col_max = col_data.max()
                                
                                # 确保所有值都是有效数字
                                if pd.notna(col_mean) and pd.notna(col_min) and pd.notna(col_max):
                                    value_columns.append({
                                        'name': col,
                                        'avg': float(col_mean),
                                        'min': float(col_min),
                                        'max': float(col_max),
                                        'non_zero_ratio': float(non_zero_ratio),
                                        'is_converted': try_conversion
                                    })
                        except Exception as e:
                            print(f"处理列 '{col}' 时出错: {str(e)}")
                
                #print(f"检测到的值列: {[col['name'] for col in value_columns]}")
                
                # 推荐分析类型
                recommended_analysis = []
                
                # 如果有日期列和值列，推荐时间序列分析
                if date_columns and value_columns:
                    recommended_analysis.append({
                        'type': 'trend',
                        'name': '趋势分析',
                        'description': '分析销售额随时间的变化趋势'
                    })
                    
                    # 检查是否有跨年数据
                    if len(date_columns) > 0:
                        try:
                            dates = pd.to_datetime(df[date_columns[0]], errors='coerce')
                            years = dates.dt.year.unique()
                            if len(years) > 1:
                                recommended_analysis.append({
                                    'type': 'year_over_year',
                                    'name': '同比分析',
                                    'description': '比较不同年份同期的销售表现'
                                })
                        except:
                            pass
                    
                    # 检查是否有跨月数据
                    if len(date_columns) > 0:
                        try:
                            dates = pd.to_datetime(df[date_columns[0]], errors='coerce')
                            year_months = dates.dt.strftime('%Y-%m').unique()
                            if len(year_months) > 1:
                                recommended_analysis.append({
                                    'type': 'month_over_month',
                                    'name': '环比分析',
                                    'description': '比较相邻月份的销售表现'
                                })
                        except:
                            pass
            
            # 确保总是返回实际行数
            actual_row_count = max(estimated_rows, row_count)
            #print(f"返回给前端的行数: {actual_row_count}")
            
            # 返回建议结果
            response_data = {
                'success': True,
                'date_columns': date_columns,
                'value_columns': value_columns,
                'recommended_analysis': recommended_analysis,
                'row_count': actual_row_count,  # 添加行数信息
                'sheet_names': sheet_names,     # 添加工作表列表
                'multiple_sheets': multiple_sheets,  # 是否有多个工作表
                'selected_sheet': selected_sheet     # 当前选择的工作表
            }
            
            #print(f"返回给前端的完整响应: {response_data}")
            return jsonify(response_data)
            
        except Exception as e:
            return jsonify({'success': False, 'message': f'分析文件时出错: {str(e)}'})
        
        finally:
            # 清理临时文件
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'处理请求时出错: {str(e)}'})

# 添加自定义分解函数
def simple_decompose(series, period):
    """简单的时间序列分解函数，替代statsmodels.tsa.seasonal.seasonal_decompose"""
    # 确保series有适当的索引
    if not isinstance(series.index, pd.DatetimeIndex):
        #print("警告: 时间序列分解需要DatetimeIndex，尝试转换")
        # 尝试创建简单的数字索引
        series = pd.Series(series.values, index=pd.RangeIndex(start=0, stop=len(series)))
    
    # 计算移动平均作为趋势
    trend = series.rolling(window=period, center=True).mean()
    
    # 去除趋势获得季节性+残差
    detrended = series - trend
    
    # 计算季节性成分
    if hasattr(series.index, 'month'):
        # 如果索引有月份属性，按月分组
        seasonal_groups = detrended.groupby(series.index.month)
    else:
        # 否则按周期位置分组
        seasonal_groups = detrended.groupby(np.arange(len(series)) % period)
    
    seasonal_means = seasonal_groups.mean()
    seasonal = np.zeros_like(series)
    
    for i in range(len(series)):
        if hasattr(series.index, 'month'):
            month = series.index[i].month
            seasonal[i] = seasonal_means.get(month, 0)
        else:
            idx = i % period
            seasonal[i] = seasonal_means.get(idx, 0)
    
    seasonal = pd.Series(seasonal, index=series.index)
    
    # 计算残差
    resid = series - trend - seasonal
    
    # 创建返回对象，模拟statsmodels的返回结构
    class DecomposeResult:
        def __init__(self, trend, seasonal, resid):
            self.trend = trend
            self.seasonal = seasonal
            self.resid = resid
    
    return DecomposeResult(trend, seasonal, resid)

# 添加大数据优化函数
def optimize_large_dataframe(df, date_column, value_column):
    """优化大型数据集以提高性能"""
    # 获取行数
    row_count = len(df)
    
    # 小数据集不需要优化
    if row_count < 100000:
        return df
        
    #print(f"优化大数据集: {row_count}行")
    
    # 1. 仅保留必要的列
    if len(df.columns) > 2:
        df = df[[date_column, value_column]].copy()
        
    # 2. 转换日期列
    df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
    
    # 3. 数据聚合与降采样
    # 按日期聚合，合并同一天的记录
    df = df.groupby(df[date_column].dt.date).agg({value_column: 'sum'}).reset_index()
    
    # 对于超大数据集进行时间降采样
    if row_count > 500000:
        date_min = df[date_column].min()
        date_max = df[date_column].max()
        date_span = (date_max - date_min).days
        
        # 确定合适的采样间隔
        if date_span > 1000:  # 超过3年数据
            # 按周聚合
            df[date_column] = pd.to_datetime(df[date_column])
            df['week'] = df[date_column].dt.isocalendar().week
            df['year'] = df[date_column].dt.isocalendar().year
            df = df.groupby(['year', 'week']).agg({
                date_column: 'first',  # 保留每周第一天
                value_column: 'sum'    # 合计销售额
            }).reset_index(drop=True)
        elif date_span > 365:  # 超过1年数据
            # 按3天聚合
            df[date_column] = pd.to_datetime(df[date_column])
            df['period'] = (df[date_column] - date_min).dt.days // 3
            df = df.groupby('period').agg({
                date_column: 'first',  # 保留每个时间段第一天
                value_column: 'sum'    # 合计销售额
            }).reset_index(drop=True)
    
    return df 
    return df 

def check_year_over_year_eligibility():
    """检查数据是否满足同比分析要求（至少包含两年的数据）"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '未上传文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '文件格式不正确，请上传Excel文件'})
        
        date_column = request.form.get('date_column')
        if not date_column:
            return jsonify({'success': False, 'message': '未指定日期列'})
            
        sheet_name = request.form.get('sheet_name', 0)  # 默认使用第一个工作表
        
        # 保存上传的文件
        temp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 
                              f"temp_{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}")
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)
        
        try:
            # 尝试将日期列转换为日期类型
            df = pd.read_excel(temp_path, sheet_name=sheet_name)
            
            # 检查日期列是否存在
            if date_column not in df.columns:
                return jsonify({'success': False, 'message': f'指定的日期列 "{date_column}" 不存在'})
            
            # 转换为日期时间
            df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
            
            # 删除无效日期
            invalid_dates = df[date_column].isna().sum()
            if invalid_dates > 0:
                print(f"警告: 删除 {invalid_dates} 行无效日期")
                df = df.dropna(subset=[date_column])
            
            # 检查是否有至少两年的数据
            if df[date_column].dt.year.nunique() >= 2:
                return jsonify({'success': True, 'has_enough_data': True})
            else:
                return jsonify({'success': True, 'has_enough_data': False})
        
        except Exception as e:
            return jsonify({'success': False, 'message': f'检查数据时出错: {str(e)}'})
    except Exception as e:
        print(f"检查同比分析条件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        # 清理临时文件
        if 'temp_path' in locals() and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass